"""XLSX rendering adapter using openpyxl."""

from __future__ import annotations

import logging
import re
import time
from pathlib import Path
from typing import Any, Dict, List, Optional

from backend.engine.domain.reports import OutputFormat
from .base import BaseRenderer, RenderContext, RenderResult

logger = logging.getLogger("neura.adapters.rendering.xlsx")


class XLSXRenderer(BaseRenderer):
    """Renderer that produces XLSX output from HTML tables."""

    @property
    def output_format(self) -> OutputFormat:
        return OutputFormat.XLSX

    def render(self, context: RenderContext) -> RenderResult:
        """Render XLSX from HTML tables."""
        start = time.perf_counter()

        try:
            self._ensure_output_dir(context.output_path)

            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            except ImportError:
                raise ImportError(
                    "openpyxl is required for XLSX rendering. "
                    "Install with: pip install openpyxl"
                )

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"

            # Extract tables from HTML and write to worksheet
            tables = self._extract_tables_from_html(context.template_html)

            if tables:
                self._write_tables_to_worksheet(ws, tables)
            else:
                # Fallback: write raw data if available
                if context.data:
                    self._write_data_to_worksheet(ws, context.data)

            # Auto-fit columns
            self._auto_fit_columns(ws)

            # Save workbook
            wb.save(str(context.output_path))

            render_time = (time.perf_counter() - start) * 1000
            return RenderResult(
                success=True,
                output_path=context.output_path,
                format=OutputFormat.XLSX,
                size_bytes=self._get_file_size(context.output_path),
                render_time_ms=render_time,
            )
        except Exception as e:
            logger.exception("xlsx_render_failed")
            return RenderResult(
                success=False,
                output_path=None,
                format=OutputFormat.XLSX,
                error=str(e),
                render_time_ms=(time.perf_counter() - start) * 1000,
            )

    def _extract_tables_from_html(self, html: str) -> List[List[List[str]]]:
        """Extract table data from HTML."""
        from html.parser import HTMLParser

        class TableParser(HTMLParser):
            def __init__(self):
                super().__init__()
                self.tables = []
                self.current_table = []
                self.current_row = []
                self.current_cell = ""
                self.in_table = False
                self.in_cell = False

            def handle_starttag(self, tag, attrs):
                if tag == "table":
                    self.in_table = True
                    self.current_table = []
                elif tag == "tr" and self.in_table:
                    self.current_row = []
                elif tag in ("td", "th") and self.in_table:
                    self.in_cell = True
                    self.current_cell = ""

            def handle_endtag(self, tag):
                if tag == "table":
                    if self.current_table:
                        self.tables.append(self.current_table)
                    self.in_table = False
                elif tag == "tr" and self.in_table:
                    if self.current_row:
                        self.current_table.append(self.current_row)
                elif tag in ("td", "th") and self.in_table:
                    self.current_row.append(self.current_cell.strip())
                    self.in_cell = False

            def handle_data(self, data):
                if self.in_cell:
                    self.current_cell += data

        parser = TableParser()
        parser.feed(html)
        return parser.tables

    def _write_tables_to_worksheet(self, ws, tables: List[List[List[str]]]) -> None:
        """Write extracted tables to worksheet."""
        from openpyxl.styles import Font, Border, Side, PatternFill

        current_row = 1

        for table_idx, table in enumerate(tables):
            if table_idx > 0:
                current_row += 2  # Gap between tables

            for row_idx, row in enumerate(table):
                for col_idx, cell_value in enumerate(row):
                    cell = ws.cell(row=current_row, column=col_idx + 1, value=cell_value)

                    # Style header row
                    if row_idx == 0:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="CCCCCC",
                            end_color="CCCCCC",
                            fill_type="solid",
                        )

                    # Add borders
                    thin_border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )
                    cell.border = thin_border

                current_row += 1

    def _write_data_to_worksheet(self, ws, data: Dict[str, Any]) -> None:
        """Write raw data dictionary to worksheet."""
        from openpyxl.styles import Font

        current_row = 1

        # Write scalars
        if "scalars" in data:
            for key, value in data["scalars"].items():
                ws.cell(row=current_row, column=1, value=key).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=str(value))
                current_row += 1
            current_row += 1

        # Write rows
        if "rows" in data and data["rows"]:
            rows = data["rows"]
            if rows:
                # Headers
                for col_idx, key in enumerate(rows[0].keys()):
                    cell = ws.cell(row=current_row, column=col_idx + 1, value=key)
                    cell.font = Font(bold=True)
                current_row += 1

                # Data rows
                for row in rows:
                    for col_idx, value in enumerate(row.values()):
                        ws.cell(row=current_row, column=col_idx + 1, value=str(value))
                    current_row += 1

    def _auto_fit_columns(self, ws) -> None:
        """Auto-fit column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def render_xlsx_from_html(
    html_path: Path,
    output_path: Path,
) -> RenderResult:
    """Convenience function to render XLSX from HTML file."""
    html_content = html_path.read_text(encoding="utf-8")

    renderer = XLSXRenderer()
    context = RenderContext(
        template_html=html_content,
        data={},
        output_format=OutputFormat.XLSX,
        output_path=output_path,
    )

    return renderer.render(context)
