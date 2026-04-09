from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

# Prefer xlsxwriter (streaming, constant memory) over openpyxl (in-memory).
_xlsxwriter = None
try:
    import xlsxwriter as _xlsxwriter  # type: ignore
except ImportError:
    _xlsxwriter = None

# Fallback: openpyxl (loads entire workbook into RAM — OOM on large reports).
_openpyxl = None
try:
    import openpyxl as _openpyxl  # type: ignore
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore
except ImportError:  # pragma: no cover
    _openpyxl = None
    Alignment = None  # type: ignore
    Border = None  # type: ignore
    Font = None  # type: ignore
    PatternFill = None  # type: ignore
    Side = None  # type: ignore
    get_column_letter = None  # type: ignore
    Table = None  # type: ignore
    TableStyleInfo = None  # type: ignore

from .html_table_parser import extract_first_table, extract_tables, extract_tables_with_header_counts

logger = logging.getLogger("neura.reports.xlsx")


# ---------------------------------------------------------------------------
# Shared HTML → rows parsing
# ---------------------------------------------------------------------------

def _table_score(table: list[list[str]]) -> int:
    if not table:
        return 0
    row_count = len(table)
    max_cols = max((len(row) for row in table), default=0)
    multi_col_rows = sum(1 for row in table if sum(1 for cell in row if cell.strip()) >= 2)
    return (multi_col_rows or row_count) * max(1, max_cols)


def _select_best_table_index(tables: list[list[list[str]]]) -> int:
    best_idx = 0
    best_score = -1
    for idx, table in enumerate(tables):
        score = _table_score(table)
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx


def _looks_like_total_row(row: list[str]) -> bool:
    for cell in row:
        if cell and cell.strip().lower() == "total":
            return True
    return False


def _parse_html_to_rows(html_text: str):
    """Parse HTML and return (rows, data_row_positions, preface_ranges, data_header_row_idx, data_thead_count)."""
    tables_with_meta = extract_tables_with_header_counts(html_text)
    tables = [t for t, _ in tables_with_meta]
    thead_counts = [c for _, c in tables_with_meta]
    rows: list[list[str]] = []
    data_row_positions: list[int] = []
    preface_ranges: list[tuple[int, int]] = []
    data_header_row_idx: int | None = None

    if tables:
        best_idx = _select_best_table_index(tables)
        data_thead_count = thead_counts[best_idx] if best_idx < len(thead_counts) else 1

        for idx, table in enumerate(tables):
            is_data_table = idx == best_idx
            if not table:
                continue
            serial_counter = 0
            table_start_idx = len(rows) + 1
            header_rows = data_thead_count if is_data_table else 1
            for row_idx_in_table, row in enumerate(table):
                clean_row = [(cell or "").strip() for cell in row]
                if is_data_table and row_idx_in_table >= header_rows:
                    if _looks_like_total_row(clean_row):
                        if clean_row:
                            clean_row[0] = ""
                    else:
                        serial_counter += 1
                        if not clean_row[0]:
                            clean_row[0] = str(serial_counter)
                elif is_data_table and row_idx_in_table == header_rows - 1:
                    data_header_row_idx = len(rows) + 1
                rows.append(clean_row)
                if is_data_table and row_idx_in_table >= header_rows:
                    data_row_positions.append(len(rows))
            table_end_idx = len(rows)
            if (not is_data_table) and table_end_idx >= table_start_idx:
                preface_ranges.append((table_start_idx, table_end_idx))
            rows.append([])

        while rows and not rows[-1]:
            rows.pop()
    else:
        rows = [[line.strip()] for line in html_text.splitlines() if line.strip()]
        if not rows:
            rows = [["Report output unavailable"]]

    # Pad all rows to the same width so preface/header rows span the full
    # sheet width in Excel (prevents narrow preface rows when data has many columns).
    max_cols = max((len(r) for r in rows if r), default=1)
    for i, row in enumerate(rows):
        if row and len(row) < max_cols:
            rows[i] = row + [""] * (max_cols - len(row))

    return rows, data_row_positions, preface_ranges, data_header_row_idx


# ---------------------------------------------------------------------------
# xlsxwriter-based export (streaming, constant memory)
# ---------------------------------------------------------------------------

def _html_file_to_xlsx_xlsxwriter(html_path: Path, output_path: Path) -> Optional[Path]:
    """Export HTML to XLSX using xlsxwriter (streaming writer, constant memory)."""
    html_text = html_path.read_text(encoding="utf-8", errors="ignore")
    rows, data_row_positions, preface_ranges, data_header_row_idx = _parse_html_to_rows(html_text)

    data_start = data_row_positions[0] if data_row_positions else None
    data_end = data_row_positions[-1] if data_row_positions else None
    data_max_cols = (
        max(len(rows[idx - 1]) for idx in data_row_positions)
        if data_row_positions
        else max((len(r) for r in rows if r), default=1)
    )
    data_max_cols = max(1, data_max_cols)
    sheet_width = max((len(r) for r in rows if r), default=data_max_cols)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = _xlsxwriter.Workbook(str(output_path), {"constant_memory": False})
    ws = wb.add_worksheet("Report")

    # Pre-define format objects (xlsxwriter requires this)
    fmt_default = wb.add_format({
        "text_wrap": True,
        "valign": "top",
        "align": "left",
    })
    fmt_bold = wb.add_format({
        "bold": True,
        "text_wrap": True,
        "valign": "top",
        "align": "left",
    })
    fmt_title = wb.add_format({
        "bold": True,
        "font_size": 14,
        "bg_color": "#BDD7EE",
        "text_wrap": True,
        "valign": "vcenter",
        "align": "center",
    })
    fmt_preface = wb.add_format({
        "bold": True,
        "font_size": 11,
        "bg_color": "#D9E1F2",
        "text_wrap": True,
        "valign": "vcenter",
        "align": "center",
    })
    fmt_data_header = wb.add_format({
        "bold": True,
        "font_color": "#FFFFFF",
        "bg_color": "#2F75B5",
        "text_wrap": True,
        "valign": "vcenter",
        "align": "center",
    })
    fmt_border = wb.add_format({
        "border": 1,
        "border_color": "#C0C0C0",
        "text_wrap": True,
        "valign": "top",
        "align": "left",
    })

    # Build set lookups for fast row classification
    preface_row_set: set[int] = set()
    first_preface_row = preface_ranges[0][0] if preface_ranges else None
    for start_idx, end_idx in preface_ranges:
        for ri in range(start_idx, end_idx + 1):
            preface_row_set.add(ri)

    # Track max column widths for auto-sizing
    col_widths: dict[int, int] = {}

    for r_idx, row in enumerate(rows, start=1):
        if not row:
            continue

        # Determine format for this row
        is_preface = r_idx in preface_row_set
        is_data_header = r_idx == data_header_row_idx
        is_title = is_preface and r_idx == first_preface_row

        if is_title:
            row_fmt = fmt_title
        elif is_preface:
            row_fmt = fmt_preface
        elif is_data_header:
            row_fmt = fmt_data_header
        elif r_idx == 1:
            row_fmt = fmt_bold
        else:
            row_fmt = fmt_border

        # For preface title/subtitle rows with only 1 non-empty cell,
        # merge across all data columns so the text spans the full width.
        non_empty = [i for i, v in enumerate(row) if v and str(v).strip()]
        if (is_title or (is_preface and len(non_empty) == 1)) and data_max_cols > 1:
            val = row[non_empty[0]] if non_empty else ""
            ws.merge_range(r_idx - 1, 0, r_idx - 1, data_max_cols - 1, val, row_fmt)
        else:
            for c_idx, value in enumerate(row):
                ws.write(r_idx - 1, c_idx, value, row_fmt)
        # Track width
        for c_idx, value in enumerate(row):
            text_len = len(str(value)) if value else 0
            old = col_widths.get(c_idx, 0)
            if text_len > old:
                col_widths[c_idx] = text_len

    # Set column widths
    for c_idx, max_len in col_widths.items():
        width = min(120, max(12, max_len + 2))
        ws.set_column(c_idx, c_idx, width)

    # Freeze panes
    if data_header_row_idx:
        freeze_row = data_header_row_idx  # 1-based → 0-based is data_header_row_idx - 1, but freeze expects row below
        ws.freeze_panes(freeze_row, 0)
    elif data_start:
        ws.freeze_panes(data_start, 0)
    elif len(rows) > 1:
        ws.freeze_panes(1, 0)

    # Autofilter on the data range
    if data_header_row_idx and data_end:
        ws.autofilter(data_header_row_idx - 1, 0, data_end - 1, data_max_cols - 1)
    elif len(rows) > 0:
        ws.autofilter(0, 0, len(rows) - 1, sheet_width - 1)

    try:
        wb.close()
    except Exception as exc:  # pragma: no cover
        logger.warning(
            "xlsx_export_save_failed",
            extra={
                "event": "xlsx_export_save_failed",
                "html_path": str(html_path),
                "xlsx_path": str(output_path),
                "error": str(exc),
                "engine": "xlsxwriter",
            },
        )
        return None

    logger.info(
        "xlsx_export_success",
        extra={
            "event": "xlsx_export_success",
            "html_path": str(html_path),
            "xlsx_path": str(output_path),
            "engine": "xlsxwriter",
        },
    )
    return output_path


# ---------------------------------------------------------------------------
# openpyxl-based export (fallback — loads everything into RAM)
# ---------------------------------------------------------------------------

def _auto_column_widths(worksheet) -> None:
    if get_column_letter is None:  # pragma: no cover
        return
    for col_idx in range(1, worksheet.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in worksheet[letter]:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            max_len = max(max_len, len(text))
        width = min(120, max(12, max_len + 2))
        worksheet.column_dimensions[letter].width = width


def _html_file_to_xlsx_openpyxl(html_path: Path, output_path: Path) -> Optional[Path]:
    """Export HTML to XLSX using openpyxl (in-memory — may OOM on large reports)."""
    html_text = html_path.read_text(encoding="utf-8", errors="ignore")
    rows, data_row_positions, preface_ranges, data_header_row_idx = _parse_html_to_rows(html_text)

    data_start = data_row_positions[0] if data_row_positions else None
    data_end = data_row_positions[-1] if data_row_positions else None
    data_max_cols = (
        max(len(rows[idx - 1]) for idx in data_row_positions)
        if data_row_positions
        else max((len(r) for r in rows if r), default=1)
    )
    data_max_cols = max(1, data_max_cols)
    sheet_width = max((len(r) for r in rows if r), default=data_max_cols)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if Alignment is not None:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True,
                )
            if Font is not None and r_idx == 1:
                cell.font = Font(bold=True)

        if len(row) == 0:
            continue

    if (
        PatternFill is not None
        and Alignment is not None
        and Font is not None
        and sheet_width > 0
        and preface_ranges
    ):
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        title_fill = PatternFill("solid", fgColor="BDD7EE")
        first_preface_row = preface_ranges[0][0]
        for start_idx, end_idx in preface_ranges:
            for row_idx in range(start_idx, end_idx + 1):
                row_data = rows[row_idx - 1] if 0 <= row_idx - 1 < len(rows) else []
                if not row_data or not any(cell for cell in row_data):
                    continue
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=sheet_width)
                cell = ws.cell(row=row_idx, column=1)
                cell_text = cell.value or ""
                is_title_row = row_idx == first_preface_row and cell_text and cell_text.upper() == cell_text
                cell.alignment = Alignment(
                    horizontal="center" if is_title_row else "left",
                    vertical="center",
                    wrap_text=True,
                )
                cell.font = Font(bold=True, size=14 if is_title_row else 11)
                cell.fill = title_fill if is_title_row else header_fill

    if (
        PatternFill is not None
        and Alignment is not None
        and Font is not None
        and data_header_row_idx is not None
        and data_max_cols > 0
    ):
        data_header_fill = PatternFill("solid", fgColor="2F75B5")
        for col_idx in range(1, data_max_cols + 1):
            cell = ws.cell(row=data_header_row_idx, column=col_idx)
            cell.fill = data_header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if Border is not None and Side is not None:
        thin = Side(style="thin", color="FFC0C0C0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        ):
            for cell in row:
                cell.border = border

    if data_header_row_idx:
        freeze_row = data_header_row_idx + 1
    else:
        freeze_row = (data_start + 1) if data_start else 2
    if ws.max_row >= freeze_row:
        ws.freeze_panes = f"A{freeze_row}"
    elif ws.max_row > 1:
        ws.freeze_panes = "A2"

    if Table is not None and TableStyleInfo is not None and ws.max_column > 0 and data_end is not None:
        table_cols = max(1, min(data_max_cols, ws.max_column))
        table_top = data_header_row_idx if data_header_row_idx is not None else data_start
        if table_top is not None and table_top <= data_end:
            ref = f"A{table_top}:{ws.cell(row=data_end, column=table_cols).coordinate}"
            table = Table(displayName="ReportTable", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
        else:
            ws.auto_filter.ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    else:
        ws.auto_filter.ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"

    _auto_column_widths(ws)

    try:
        wb.save(output_path)
    except Exception as exc:  # pragma: no cover
        logger.warning(
            "xlsx_export_save_failed",
            extra={
                "event": "xlsx_export_save_failed",
                "html_path": str(html_path),
                "xlsx_path": str(output_path),
                "error": str(exc),
                "engine": "openpyxl",
            },
        )
        return None

    logger.info(
        "xlsx_export_success",
        extra={
            "event": "xlsx_export_success",
            "html_path": str(html_path),
            "xlsx_path": str(output_path),
            "engine": "openpyxl",
        },
    )
    return output_path


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def html_file_to_xlsx(html_path: Path, output_path: Path) -> Optional[Path]:
    """Convert an HTML report file to XLSX.

    Uses xlsxwriter (streaming, constant memory) when available.
    Falls back to openpyxl if xlsxwriter is not installed.
    """
    if _xlsxwriter is not None:
        return _html_file_to_xlsx_xlsxwriter(html_path, output_path)
    if _openpyxl is not None:
        logger.info("xlsx_export_using_openpyxl_fallback")
        return _html_file_to_xlsx_openpyxl(html_path, output_path)
    logger.warning(
        "xlsx_export_unavailable",
        extra={
            "event": "xlsx_export_unavailable",
            "reason": "neither xlsxwriter nor openpyxl installed",
            "html_path": str(html_path),
        },
    )
    return None
