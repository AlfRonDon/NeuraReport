# mypy: ignore-errors
"""
Enhanced Document Extraction Module.

Combines traditional extraction methods with AI-powered understanding:
- PDF extraction with layout analysis
- Excel/spreadsheet parsing
- VLM-enhanced table extraction
- Automatic schema detection
- Multi-format support
"""
from __future__ import annotations

import json
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

logger = logging.getLogger("neura.llm.document_extractor")


@dataclass
class ExtractedTable:
    """Extracted table from a document."""
    id: str
    title: Optional[str]
    headers: List[str]
    rows: List[List[Any]]
    metadata: Dict[str, Any] = field(default_factory=dict)
    confidence: float = 1.0


@dataclass
class ExtractedContent:
    """Complete extracted content from a document."""
    text: str
    tables: List[ExtractedTable]
    metadata: Dict[str, Any]
    structure: Dict[str, Any]
    warnings: List[str] = field(default_factory=list)


@dataclass
class FieldSchema:
    """Schema for an extracted field."""
    name: str
    data_type: str  # text, numeric, datetime, boolean
    sample_values: List[Any]
    nullable: bool = True
    description: Optional[str] = None


class EnhancedDocumentExtractor:
    """
    Enhanced document extractor with AI-powered understanding.

    Combines:
    - Traditional PDF/Excel extraction
    - Layout analysis
    - VLM for complex documents
    - Automatic schema inference
    """

    def __init__(
        self,
        use_vlm: bool = True,
        use_ocr: bool = True,
        max_pages: int = 50,
        max_tables: int = 100,
    ):
        self.use_vlm = use_vlm
        self.use_ocr = use_ocr
        self.max_pages = max_pages
        self.max_tables = max_tables

        # Lazy load VLM
        self._vlm = None

    @property
    def vlm(self):
        """Get VLM instance (lazy loaded)."""
        if self._vlm is None and self.use_vlm:
            try:
                from .vision import get_vlm
                self._vlm = get_vlm()
            except Exception as e:
                logger.warning(f"VLM not available: {e}")
                self._vlm = None
        return self._vlm

    def extract(
        self,
        file_path: Union[str, Path],
        extraction_mode: str = "auto",
    ) -> ExtractedContent:
        """
        Extract content from a document.

        Args:
            file_path: Path to the document
            extraction_mode: auto, text_only, tables_only, comprehensive

        Returns:
            ExtractedContent with all extracted data
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        suffix = file_path.suffix.lower()

        if suffix == ".pdf":
            return self._extract_pdf(file_path, extraction_mode)
        elif suffix in (".xlsx", ".xls", ".xlsm"):
            return self._extract_excel(file_path, extraction_mode)
        elif suffix == ".csv":
            return self._extract_csv(file_path)
        elif suffix in (".png", ".jpg", ".jpeg", ".tiff", ".bmp"):
            return self._extract_image(file_path, extraction_mode)
        else:
            raise ValueError(f"Unsupported file format: {suffix}")

    def _extract_pdf(
        self,
        file_path: Path,
        mode: str,
    ) -> ExtractedContent:
        """Extract content from PDF."""
        try:
            import fitz  # PyMuPDF
        except ImportError:
            raise RuntimeError("PyMuPDF is required for PDF extraction. Install with: pip install pymupdf")

        doc = fitz.open(file_path)
        warnings = []

        if doc.page_count > self.max_pages:
            warnings.append(f"PDF has {doc.page_count} pages, only processing first {self.max_pages}")

        text_content = []
        tables: List[ExtractedTable] = []
        metadata = {
            "filename": file_path.name,
            "page_count": doc.page_count,
            "format": "pdf",
        }

        for page_num, page in enumerate(doc):
            if page_num >= self.max_pages:
                break

            # Extract text
            page_text = page.get_text("text")
            text_content.append(f"--- Page {page_num + 1} ---\n{page_text}")

            # Extract tables using built-in method
            page_tables = self._extract_pdf_tables(page, page_num)
            tables.extend(page_tables)

            # Use VLM for complex layouts if enabled
            if self.vlm and mode in ("auto", "comprehensive"):
                try:
                    pix = page.get_pixmap(dpi=150)
                    img_bytes = pix.tobytes("png")
                    vlm_result = self.vlm.extract_tables(img_bytes)

                    # Merge VLM-extracted tables
                    for i, vt in enumerate(vlm_result.tables):
                        table_id = f"page{page_num + 1}_vlm_{i + 1}"
                        if not self._table_exists(tables, vt.get("headers", [])):
                            tables.append(ExtractedTable(
                                id=table_id,
                                title=vt.get("title"),
                                headers=vt.get("headers", []),
                                rows=vt.get("rows", []),
                                metadata={"source": "vlm", "page": page_num + 1},
                                confidence=vlm_result.confidence,
                            ))
                except Exception as e:
                    logger.warning(f"VLM extraction failed for page {page_num + 1}: {e}")

            if len(tables) >= self.max_tables:
                warnings.append(f"Table limit ({self.max_tables}) reached")
                break

        doc.close()

        # Analyze document structure
        structure = self._analyze_structure("\n\n".join(text_content))

        return ExtractedContent(
            text="\n\n".join(text_content),
            tables=tables,
            metadata=metadata,
            structure=structure,
            warnings=warnings,
        )

    def _extract_pdf_tables(
        self,
        page,
        page_num: int,
    ) -> List[ExtractedTable]:
        """Extract tables from a PDF page."""
        tables = []

        try:
            # Use PyMuPDF's table detection
            page_tables = page.find_tables()

            for i, table in enumerate(page_tables):
                if table.row_count == 0:
                    continue

                # Extract table data
                data = table.extract()
                if not data or len(data) < 2:
                    continue

                # First row as headers
                headers = [str(cell or "").strip() for cell in data[0]]

                # Remaining rows as data
                rows = []
                for row in data[1:]:
                    normalized_row = []
                    for j, cell in enumerate(row):
                        if j < len(headers):
                            normalized_row.append(str(cell or "").strip())
                    # Pad row to match headers
                    while len(normalized_row) < len(headers):
                        normalized_row.append("")
                    rows.append(normalized_row)

                table_id = f"page{page_num + 1}_table_{i + 1}"
                tables.append(ExtractedTable(
                    id=table_id,
                    title=None,
                    headers=headers,
                    rows=rows,
                    metadata={"source": "pymupdf", "page": page_num + 1},
                    confidence=0.9,
                ))

        except Exception as e:
            logger.warning(f"Table extraction failed for page {page_num + 1}: {e}")

        return tables

    def _extract_excel(
        self,
        file_path: Path,
        mode: str,
    ) -> ExtractedContent:
        """Extract content from Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required for Excel extraction. Install with: pip install openpyxl")

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        warnings = []

        tables: List[ExtractedTable] = []
        text_parts = []
        metadata = {
            "filename": file_path.name,
            "sheet_count": len(workbook.sheetnames),
            "format": "excel",
            "sheets": workbook.sheetnames,
        }

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            if sheet.max_row == 0 or sheet.max_column == 0:
                continue

            # Convert sheet to table
            headers = []
            rows = []

            for row_num, row in enumerate(sheet.iter_rows(values_only=True)):
                # Clean row
                cleaned_row = [str(cell) if cell is not None else "" for cell in row]

                if row_num == 0:
                    headers = cleaned_row
                else:
                    # Skip empty rows
                    if any(cell.strip() for cell in cleaned_row):
                        # Normalize row length
                        while len(cleaned_row) < len(headers):
                            cleaned_row.append("")
                        rows.append(cleaned_row[:len(headers)])

            if headers and rows:
                tables.append(ExtractedTable(
                    id=f"sheet_{sheet_name}",
                    title=sheet_name,
                    headers=headers,
                    rows=rows,
                    metadata={"source": "openpyxl", "sheet": sheet_name},
                    confidence=1.0,
                ))

            # Build text representation
            text_parts.append(f"=== Sheet: {sheet_name} ===")
            text_parts.append("\t".join(headers))
            for row in rows[:10]:  # First 10 rows for text preview
                text_parts.append("\t".join(row))
            if len(rows) > 10:
                text_parts.append(f"... ({len(rows) - 10} more rows)")

        workbook.close()

        return ExtractedContent(
            text="\n".join(text_parts),
            tables=tables,
            metadata=metadata,
            structure={"type": "spreadsheet", "sheets": workbook.sheetnames},
            warnings=warnings,
        )

    def _extract_csv(self, file_path: Path) -> ExtractedContent:
        """Extract content from CSV file."""
        import csv

        tables = []
        warnings = []

        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            # Detect delimiter
            sample = f.read(4096)
            f.seek(0)

            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel

            reader = csv.reader(f, dialect)
            rows_list = list(reader)

        if not rows_list:
            return ExtractedContent(
                text="",
                tables=[],
                metadata={"filename": file_path.name, "format": "csv"},
                structure={"type": "empty"},
                warnings=["CSV file is empty"],
            )

        headers = rows_list[0] if rows_list else []
        rows = rows_list[1:] if len(rows_list) > 1 else []

        # Normalize row lengths
        max_cols = max(len(row) for row in rows_list) if rows_list else 0
        while len(headers) < max_cols:
            headers.append(f"Column_{len(headers) + 1}")

        normalized_rows = []
        for row in rows:
            normalized = list(row)
            while len(normalized) < len(headers):
                normalized.append("")
            normalized_rows.append(normalized[:len(headers)])

        tables.append(ExtractedTable(
            id="csv_table",
            title=file_path.stem,
            headers=headers,
            rows=normalized_rows,
            metadata={"source": "csv"},
            confidence=1.0,
        ))

        # Text representation
        text_parts = ["\t".join(headers)]
        for row in normalized_rows[:20]:
            text_parts.append("\t".join(row))
        if len(normalized_rows) > 20:
            text_parts.append(f"... ({len(normalized_rows) - 20} more rows)")

        return ExtractedContent(
            text="\n".join(text_parts),
            tables=tables,
            metadata={
                "filename": file_path.name,
                "format": "csv",
                "row_count": len(normalized_rows),
                "column_count": len(headers),
            },
            structure={"type": "tabular"},
            warnings=warnings,
        )

    def _extract_image(
        self,
        file_path: Path,
        mode: str,
    ) -> ExtractedContent:
        """Extract content from image using VLM."""
        if not self.vlm:
            raise RuntimeError("VLM is required for image extraction but not available")

        # Use VLM for comprehensive extraction
        result = self.vlm.analyze_document(file_path, analysis_type=mode)

        tables = []
        for i, table_data in enumerate(result.tables):
            tables.append(ExtractedTable(
                id=f"image_table_{i + 1}",
                title=table_data.get("title"),
                headers=table_data.get("headers", []),
                rows=table_data.get("rows", []),
                metadata={"source": "vlm"},
                confidence=0.8,
            ))

        return ExtractedContent(
            text=result.text_content,
            tables=tables,
            metadata={
                "filename": file_path.name,
                "format": "image",
                **result.metadata,
            },
            structure=result.structure,
            warnings=[],
        )

    def _analyze_structure(self, text: str) -> Dict[str, Any]:
        """Analyze document structure from text."""
        structure = {
            "type": "document",
            "has_headers": False,
            "has_lists": False,
            "has_tables": False,
            "sections": [],
        }

        # Detect headers (lines that look like titles)
        header_pattern = re.compile(r'^[A-Z][A-Za-z\s]+:?\s*$', re.MULTILINE)
        headers = header_pattern.findall(text)
        structure["has_headers"] = len(headers) > 0
        structure["sections"] = [h.strip().rstrip(':') for h in headers[:10]]

        # Detect lists
        list_pattern = re.compile(r'^[\s]*[-•*]\s+.+$', re.MULTILINE)
        structure["has_lists"] = bool(list_pattern.search(text))

        # Detect table-like content
        table_pattern = re.compile(r'\|.+\|', re.MULTILINE)
        structure["has_tables"] = bool(table_pattern.search(text))

        return structure

    def _table_exists(
        self,
        tables: List[ExtractedTable],
        headers: List[str],
    ) -> bool:
        """Check if a table with similar headers already exists."""
        if not headers:
            return True

        for table in tables:
            if len(table.headers) == len(headers):
                # Check if headers match (case-insensitive)
                if all(
                    h1.lower().strip() == h2.lower().strip()
                    for h1, h2 in zip(table.headers, headers)
                ):
                    return True
        return False

    def infer_schema(
        self,
        table: ExtractedTable,
    ) -> List[FieldSchema]:
        """Infer schema for a table's columns."""
        schemas = []

        for col_idx, header in enumerate(table.headers):
            # Collect sample values
            sample_values = []
            for row in table.rows[:100]:
                if col_idx < len(row) and row[col_idx]:
                    sample_values.append(row[col_idx])

            # Infer data type
            data_type = self._infer_column_type(sample_values)

            schemas.append(FieldSchema(
                name=header,
                data_type=data_type,
                sample_values=sample_values[:5],
                nullable=any(not v for v in sample_values),
            ))

        return schemas

    def _infer_column_type(self, values: List[Any]) -> str:
        """Infer the data type of a column from sample values."""
        if not values:
            return "text"

        # Count type matches
        numeric_count = 0
        date_count = 0
        bool_count = 0

        date_patterns = [
            r'^\d{4}-\d{2}-\d{2}$',
            r'^\d{2}/\d{2}/\d{4}$',
            r'^\d{2}-\d{2}-\d{4}$',
            r'^\d{1,2}/\d{1,2}/\d{2,4}$',
        ]

        for value in values:
            value_str = str(value).strip()

            if not value_str:
                continue

            # Check numeric
            try:
                cleaned = re.sub(r'[$,% ]', '', value_str)
                float(cleaned)
                numeric_count += 1
                continue
            except (ValueError, TypeError):
                pass

            # Check date
            for pattern in date_patterns:
                if re.match(pattern, value_str):
                    date_count += 1
                    break

            # Check boolean
            if value_str.lower() in ('true', 'false', 'yes', 'no', '1', '0'):
                bool_count += 1

        total = len([v for v in values if str(v).strip()])
        if total == 0:
            return "text"

        # Require 70% match for type classification
        threshold = 0.7

        if date_count / total >= threshold:
            return "datetime"
        if numeric_count / total >= threshold:
            return "numeric"
        if bool_count / total >= threshold:
            return "boolean"

        return "text"


# Convenience functions

def extract_document(
    file_path: Union[str, Path],
    use_vlm: bool = True,
) -> ExtractedContent:
    """Quick function to extract content from a document."""
    extractor = EnhancedDocumentExtractor(use_vlm=use_vlm)
    return extractor.extract(file_path)


def extract_tables(
    file_path: Union[str, Path],
) -> List[ExtractedTable]:
    """Quick function to extract tables from a document."""
    extractor = EnhancedDocumentExtractor(use_vlm=True)
    result = extractor.extract(file_path, extraction_mode="tables_only")
    return result.tables
