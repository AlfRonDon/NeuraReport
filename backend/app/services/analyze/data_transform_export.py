# mypy: ignore-errors
"""
Data Transformation & Export Service - Smart data cleaning and multi-format export.

Features:
7.1 Smart Data Cleaning
7.2 Export Options (Excel, PDF, CSV, JSON, etc.)
7.3 Template Integration
"""
from __future__ import annotations

import csv
import io
import json
import logging
import math
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

from backend.app.schemas.analyze.enhanced_analysis import (
    DataQualityReport,
    DataTransformation,
    EnhancedAnalysisResult,
    EnhancedChartSpec,
    EnhancedExtractedTable,
    ExportConfiguration,
    ExportFormat,
)

logger = logging.getLogger("neura.analyze.export")


# =============================================================================
# DATA QUALITY ASSESSMENT
# =============================================================================

def assess_data_quality(tables: List[EnhancedExtractedTable]) -> DataQualityReport:
    """Assess data quality across all tables."""
    total_rows = sum(t.row_count for t in tables)
    total_columns = sum(len(t.headers) for t in tables)

    missing_values: Dict[str, int] = {}
    missing_percentage: Dict[str, float] = {}
    unique_values_per_column: Dict[str, int] = {}
    invalid_values: Dict[str, List[Any]] = {}
    type_mismatches: Dict[str, List[int]] = {}
    format_inconsistencies: Dict[str, List[str]] = {}
    outliers_detected: Dict[str, List[int]] = {}

    duplicate_rows = 0

    for table in tables:
        # Check for duplicate rows
        seen_rows = set()
        for row in table.rows:
            row_key = tuple(str(v) for v in row)
            if row_key in seen_rows:
                duplicate_rows += 1
            seen_rows.add(row_key)

        # Analyze each column
        for col_idx, (header, dtype) in enumerate(zip(table.headers, table.data_types)):
            col_key = f"{table.id}.{header}"

            # Collect column values
            values = []
            missing_count = 0
            unique_set = set()
            numeric_values = []

            for row_idx, row in enumerate(table.rows):
                if col_idx < len(row):
                    val = row[col_idx]
                    values.append((row_idx, val))

                    # Check missing
                    if val is None or str(val).strip() == "" or str(val).lower() in ("null", "n/a", "na", "-"):
                        missing_count += 1
                    else:
                        unique_set.add(str(val))

                        # Type validation
                        if dtype == "numeric":
                            try:
                                cleaned = re.sub(r'[$,% ]', '', str(val))
                                num_val = float(cleaned)
                                numeric_values.append((row_idx, num_val))
                            except (ValueError, TypeError):
                                if col_key not in type_mismatches:
                                    type_mismatches[col_key] = []
                                type_mismatches[col_key].append(row_idx)

            # Record missing values
            if missing_count > 0:
                missing_values[col_key] = missing_count
                missing_percentage[col_key] = round(missing_count / len(table.rows) * 100, 2)

            unique_values_per_column[col_key] = len(unique_set)

            # Detect outliers for numeric columns
            if numeric_values and len(numeric_values) >= 5:
                vals = [v for _, v in numeric_values]
                mean = sum(vals) / len(vals)
                std = math.sqrt(sum((v - mean) ** 2 for v in vals) / len(vals))

                if std > 0:
                    outlier_indices = []
                    for row_idx, val in numeric_values:
                        if abs((val - mean) / std) > 3:
                            outlier_indices.append(row_idx)

                    if outlier_indices:
                        outliers_detected[col_key] = outlier_indices

            # Check format inconsistencies for text columns
            if dtype == "text" and unique_set:
                # Check for mixed case patterns
                patterns_found = set()
                for val in list(unique_set)[:50]:
                    if val.isupper():
                        patterns_found.add("UPPERCASE")
                    elif val.islower():
                        patterns_found.add("lowercase")
                    elif val.istitle():
                        patterns_found.add("Title Case")
                    else:
                        patterns_found.add("Mixed case")

                if len(patterns_found) > 1:
                    format_inconsistencies[col_key] = list(patterns_found)

    # Calculate quality score
    total_issues = (
        sum(missing_values.values()) +
        duplicate_rows * 2 +
        sum(len(v) for v in type_mismatches.values()) +
        sum(len(v) for v in outliers_detected.values())
    )

    max_possible_issues = total_rows * total_columns if total_rows and total_columns else 1
    quality_score = max(0, 1 - (total_issues / max_possible_issues))

    # Generate recommendations
    recommendations = []
    if missing_values:
        high_missing = [k for k, v in missing_percentage.items() if v > 20]
        if high_missing:
            recommendations.append(f"High missing data in columns: {', '.join(high_missing[:5])}")
    if duplicate_rows > 0:
        recommendations.append(f"Found {duplicate_rows} duplicate rows - consider deduplication")
    if type_mismatches:
        recommendations.append(f"Type mismatches detected in {len(type_mismatches)} columns")
    if outliers_detected:
        recommendations.append(f"Outliers detected in {len(outliers_detected)} numeric columns")

    return DataQualityReport(
        total_rows=total_rows,
        total_columns=total_columns,
        missing_values=missing_values,
        missing_percentage=missing_percentage,
        duplicate_rows=duplicate_rows,
        unique_values_per_column=unique_values_per_column,
        invalid_values=invalid_values,
        type_mismatches=type_mismatches,
        format_inconsistencies=format_inconsistencies,
        outliers_detected=outliers_detected,
        quality_score=round(quality_score, 3),
        recommendations=recommendations,
    )


# =============================================================================
# DATA CLEANING & TRANSFORMATION
# =============================================================================

def clean_table(
    table: EnhancedExtractedTable,
    operations: List[str] = None,
) -> EnhancedExtractedTable:
    """Apply cleaning operations to a table."""
    if operations is None:
        operations = ["trim", "normalize_case", "fill_missing"]

    cleaned_rows = [list(row) for row in table.rows]  # Deep copy

    for col_idx, (header, dtype) in enumerate(zip(table.headers, table.data_types)):
        for row_idx, row in enumerate(cleaned_rows):
            if col_idx >= len(row):
                continue

            val = row[col_idx]

            if "trim" in operations:
                if isinstance(val, str):
                    val = val.strip()

            if "normalize_case" in operations:
                if isinstance(val, str) and dtype == "text":
                    val = val.title()

            if "fill_missing" in operations:
                if val is None or str(val).strip() == "" or str(val).lower() in ("null", "n/a", "na"):
                    if dtype == "numeric":
                        # Fill with column median
                        col_values = []
                        for r in table.rows:
                            if col_idx < len(r):
                                try:
                                    col_values.append(float(str(r[col_idx]).replace(",", "").replace("$", "")))
                                except (ValueError, TypeError):
                                    pass
                        if col_values:
                            sorted_vals = sorted(col_values)
                            val = sorted_vals[len(sorted_vals) // 2]
                        else:
                            val = 0
                    else:
                        val = ""

            if "normalize_numbers" in operations:
                if dtype == "numeric":
                    try:
                        cleaned = re.sub(r'[$,% ]', '', str(val))
                        val = float(cleaned)
                    except (ValueError, TypeError):
                        pass

            cleaned_rows[row_idx][col_idx] = val

    return EnhancedExtractedTable(
        id=table.id,
        title=table.title,
        headers=table.headers,
        rows=cleaned_rows,
        data_types=table.data_types,
        source_page=table.source_page,
        source_sheet=table.source_sheet,
        confidence=table.confidence,
        row_count=len(cleaned_rows),
        column_count=len(table.headers),
        has_totals_row=table.has_totals_row,
        has_header_row=table.has_header_row,
        statistics=table.statistics,
    )


def apply_transformation(
    table: EnhancedExtractedTable,
    transformation: DataTransformation,
) -> EnhancedExtractedTable:
    """Apply a single transformation to a table."""
    if transformation.operation == "clean":
        return clean_table(table, transformation.parameters.get("operations"))

    elif transformation.operation == "normalize":
        # Normalize numeric columns to 0-1 range
        for col_name in transformation.source_columns:
            if col_name not in table.headers:
                continue
            col_idx = table.headers.index(col_name)

            values = []
            for row in table.rows:
                if col_idx < len(row):
                    try:
                        values.append(float(str(row[col_idx]).replace(",", "")))
                    except (ValueError, TypeError):
                        values.append(None)

            valid_values = [v for v in values if v is not None]
            if not valid_values:
                continue

            min_val = min(valid_values)
            max_val = max(valid_values)
            range_val = max_val - min_val if max_val != min_val else 1

            for row_idx, row in enumerate(table.rows):
                if col_idx < len(row) and values[row_idx] is not None:
                    row[col_idx] = round((values[row_idx] - min_val) / range_val, 4)

    elif transformation.operation == "aggregate":
        # Aggregate by group column
        group_col = transformation.parameters.get("group_by")
        agg_func = transformation.parameters.get("function", "sum")

        if group_col not in table.headers:
            return table

        group_idx = table.headers.index(group_col)
        value_cols = transformation.source_columns

        groups: Dict[str, Dict[str, List[float]]] = {}
        for row in table.rows:
            if group_idx >= len(row):
                continue
            group_key = str(row[group_idx])

            if group_key not in groups:
                groups[group_key] = {col: [] for col in value_cols}

            for col in value_cols:
                if col in table.headers:
                    col_idx = table.headers.index(col)
                    if col_idx < len(row):
                        try:
                            groups[group_key][col].append(float(str(row[col_idx]).replace(",", "")))
                        except (ValueError, TypeError):
                            pass

        # Build aggregated rows
        new_headers = [group_col] + value_cols
        new_rows = []

        for group_key, values_dict in groups.items():
            new_row = [group_key]
            for col in value_cols:
                vals = values_dict.get(col, [])
                if not vals:
                    new_row.append(0)
                elif agg_func == "sum":
                    new_row.append(sum(vals))
                elif agg_func == "mean":
                    new_row.append(sum(vals) / len(vals))
                elif agg_func == "count":
                    new_row.append(len(vals))
                elif agg_func == "min":
                    new_row.append(min(vals))
                elif agg_func == "max":
                    new_row.append(max(vals))
                else:
                    new_row.append(sum(vals))

            new_rows.append(new_row)

        return EnhancedExtractedTable(
            id=f"{table.id}_agg",
            title=f"{table.title or table.id} (Aggregated)",
            headers=new_headers,
            rows=new_rows,
            data_types=["text"] + ["numeric"] * len(value_cols),
            row_count=len(new_rows),
            column_count=len(new_headers),
        )

    return table


# =============================================================================
# EXPORT FUNCTIONS
# =============================================================================

def export_to_csv(
    tables: List[EnhancedExtractedTable],
    include_headers: bool = True,
) -> str:
    """Export tables to CSV format."""
    output = io.StringIO()
    writer = csv.writer(output)

    for table in tables:
        if include_headers:
            writer.writerow([f"# Table: {table.title or table.id}"])
            writer.writerow(table.headers)

        for row in table.rows:
            writer.writerow(row)

        writer.writerow([])  # Empty row between tables

    return output.getvalue()


def export_to_json(
    result: EnhancedAnalysisResult,
    include_raw_data: bool = True,
) -> str:
    """Export analysis result to JSON."""
    data = result.model_dump()

    if not include_raw_data:
        # Remove large data arrays
        for table in data.get("tables", []):
            table["rows"] = table["rows"][:10]  # Only first 10 rows

    return json.dumps(data, indent=2, default=str)


def export_to_markdown(result: EnhancedAnalysisResult) -> str:
    """Export analysis result to Markdown format."""
    lines = []

    # Title
    lines.append(f"# Analysis Report: {result.document_name}")
    lines.append(f"\n*Generated: {result.created_at.strftime('%Y-%m-%d %H:%M')}*\n")

    # Executive Summary
    if "executive" in result.summaries:
        lines.append("## Executive Summary\n")
        lines.append(result.summaries["executive"].content)
        lines.append("")

    # Key Metrics
    if result.metrics:
        lines.append("## Key Metrics\n")
        for metric in result.metrics[:10]:
            change_str = f" ({metric.change:+.1f}%)" if metric.change else ""
            lines.append(f"- **{metric.name}**: {metric.raw_value}{change_str}")
        lines.append("")

    # Tables
    if result.tables:
        lines.append("## Data Tables\n")
        for table in result.tables[:5]:
            lines.append(f"### {table.title or table.id}\n")
            # Header
            lines.append("| " + " | ".join(table.headers) + " |")
            lines.append("| " + " | ".join(["---"] * len(table.headers)) + " |")
            # Rows (limit to 10)
            for row in table.rows[:10]:
                lines.append("| " + " | ".join(str(v) for v in row) + " |")
            if len(table.rows) > 10:
                lines.append(f"\n*...and {len(table.rows) - 10} more rows*\n")
            lines.append("")

    # Insights
    if result.insights:
        lines.append("## Key Insights\n")
        for insight in result.insights:
            lines.append(f"### {insight.title}")
            lines.append(f"\n{insight.description}\n")
            if insight.suggested_actions:
                lines.append("**Suggested Actions:**")
                for action in insight.suggested_actions:
                    lines.append(f"- {action}")
            lines.append("")

    # Risks
    if result.risks:
        lines.append("## Risks Identified\n")
        for risk in result.risks:
            lines.append(f"- **{risk.title}** ({risk.risk_level.value}): {risk.description}")
        lines.append("")

    # Opportunities
    if result.opportunities:
        lines.append("## Opportunities\n")
        for opp in result.opportunities:
            lines.append(f"- **{opp.title}**: {opp.description}")
        lines.append("")

    return "\n".join(lines)


def export_to_html(result: EnhancedAnalysisResult) -> str:
    """Export analysis result to HTML format."""
    html_parts = ["""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Analysis Report</title>
    <style>
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 40px; background: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 40px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        h1 { color: #1a1a2e; border-bottom: 3px solid #4f46e5; padding-bottom: 10px; }
        h2 { color: #4f46e5; margin-top: 30px; }
        h3 { color: #374151; }
        .metric { display: inline-block; background: #f0f9ff; padding: 15px 25px; margin: 5px; border-radius: 8px; border-left: 4px solid #4f46e5; }
        .metric-value { font-size: 24px; font-weight: bold; color: #1a1a2e; }
        .metric-name { color: #6b7280; font-size: 14px; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #e5e7eb; }
        th { background: #f9fafb; font-weight: 600; }
        tr:hover { background: #f9fafb; }
        .insight { background: #fef3c7; padding: 15px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #f59e0b; }
        .risk { background: #fee2e2; padding: 15px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #ef4444; }
        .opportunity { background: #d1fae5; padding: 15px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #10b981; }
        .badge { display: inline-block; padding: 4px 8px; border-radius: 4px; font-size: 12px; font-weight: 600; }
        .badge-high { background: #fee2e2; color: #dc2626; }
        .badge-medium { background: #fef3c7; color: #d97706; }
        .badge-low { background: #d1fae5; color: #059669; }
    </style>
</head>
<body>
<div class="container">
"""]

    # Title
    html_parts.append(f"<h1>Analysis Report: {result.document_name}</h1>")
    html_parts.append(f"<p style='color:#6b7280;'>Generated: {result.created_at.strftime('%Y-%m-%d %H:%M')}</p>")

    # Key Metrics
    if result.metrics:
        html_parts.append("<h2>Key Metrics</h2><div>")
        for metric in result.metrics[:8]:
            html_parts.append(f"""<div class="metric">
                <div class="metric-value">{metric.raw_value}</div>
                <div class="metric-name">{metric.name}</div>
            </div>""")
        html_parts.append("</div>")

    # Executive Summary
    if "executive" in result.summaries:
        html_parts.append("<h2>Executive Summary</h2>")
        html_parts.append(f"<p>{result.summaries['executive'].content}</p>")

    # Tables
    if result.tables:
        html_parts.append("<h2>Data Tables</h2>")
        for table in result.tables[:3]:
            html_parts.append(f"<h3>{table.title or table.id}</h3>")
            html_parts.append("<table><thead><tr>")
            for header in table.headers:
                html_parts.append(f"<th>{header}</th>")
            html_parts.append("</tr></thead><tbody>")
            for row in table.rows[:15]:
                html_parts.append("<tr>")
                for val in row:
                    html_parts.append(f"<td>{val}</td>")
                html_parts.append("</tr>")
            html_parts.append("</tbody></table>")
            if len(table.rows) > 15:
                html_parts.append(f"<p style='color:#6b7280;'>...and {len(table.rows) - 15} more rows</p>")

    # Insights
    if result.insights:
        html_parts.append("<h2>Key Insights</h2>")
        for insight in result.insights:
            html_parts.append(f"""<div class="insight">
                <strong>{insight.title}</strong>
                <span class="badge badge-{insight.priority.value}">{insight.priority.value}</span>
                <p>{insight.description}</p>
            </div>""")

    # Risks
    if result.risks:
        html_parts.append("<h2>Risks Identified</h2>")
        for risk in result.risks:
            html_parts.append(f"""<div class="risk">
                <strong>{risk.title}</strong>
                <span class="badge badge-{risk.risk_level.value}">{risk.risk_level.value}</span>
                <p>{risk.description}</p>
            </div>""")

    # Opportunities
    if result.opportunities:
        html_parts.append("<h2>Opportunities</h2>")
        for opp in result.opportunities:
            html_parts.append(f"""<div class="opportunity">
                <strong>{opp.title}</strong>
                <p>{opp.description}</p>
            </div>""")

    html_parts.append("</div></body></html>")

    return "\n".join(html_parts)


async def export_to_excel(
    result: EnhancedAnalysisResult,
    include_charts: bool = True,
) -> bytes:
    """Export analysis result to Excel format."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel export")

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Analysis Report"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A3"] = "Document:"
    ws_summary["B3"] = result.document_name
    ws_summary["A4"] = "Generated:"
    ws_summary["B4"] = result.created_at.strftime("%Y-%m-%d %H:%M")
    ws_summary["A5"] = "Tables Found:"
    ws_summary["B5"] = result.total_tables
    ws_summary["A6"] = "Metrics Extracted:"
    ws_summary["B6"] = result.total_metrics

    # Metrics Sheet
    if result.metrics:
        ws_metrics = wb.create_sheet("Metrics")
        headers = ["Name", "Value", "Type", "Period", "Change (%)", "Context"]
        for col, header in enumerate(headers, 1):
            cell = ws_metrics.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row, metric in enumerate(result.metrics, 2):
            ws_metrics.cell(row=row, column=1, value=metric.name)
            ws_metrics.cell(row=row, column=2, value=metric.raw_value)
            ws_metrics.cell(row=row, column=3, value=metric.metric_type.value)
            ws_metrics.cell(row=row, column=4, value=metric.period or "")
            ws_metrics.cell(row=row, column=5, value=metric.change or "")
            ws_metrics.cell(row=row, column=6, value=metric.context or "")

    # Data Tables
    for table in result.tables:
        # Sanitize sheet name
        sheet_name = (table.title or table.id)[:30].replace("/", "-").replace("\\", "-")
        try:
            ws_data = wb.create_sheet(sheet_name)
        except Exception:
            ws_data = wb.create_sheet(f"Table_{table.id[:20]}")

        # Headers
        for col, header in enumerate(table.headers, 1):
            cell = ws_data.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws_data.column_dimensions[get_column_letter(col)].width = max(12, len(header) + 2)

        # Data
        for row_idx, row in enumerate(table.rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws_data.cell(row=row_idx, column=col_idx, value=val)

    # Insights Sheet
    if result.insights:
        ws_insights = wb.create_sheet("Insights")
        headers = ["Type", "Priority", "Title", "Description", "Actions"]
        for col, header in enumerate(headers, 1):
            cell = ws_insights.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row, insight in enumerate(result.insights, 2):
            ws_insights.cell(row=row, column=1, value=insight.type)
            ws_insights.cell(row=row, column=2, value=insight.priority.value)
            ws_insights.cell(row=row, column=3, value=insight.title)
            ws_insights.cell(row=row, column=4, value=insight.description)
            ws_insights.cell(row=row, column=5, value="; ".join(insight.suggested_actions))

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


async def export_to_pdf(result: EnhancedAnalysisResult) -> bytes:
    """Export analysis result to PDF format."""
    html_content = export_to_html(result)

    # Try WeasyPrint first (HTML -> PDF)
    try:
        from weasyprint import HTML
        return HTML(string=html_content).write_pdf()
    except ImportError:
        pass
    except Exception as exc:
        logger.warning(f"WeasyPrint PDF export failed: {exc}")

    # Fallback to ReportLab if available
    try:
        import re
        from io import BytesIO
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas

        # Strip HTML tags for a basic text-only PDF
        text = re.sub(r"<[^>]+>", "", html_content)
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        y = height - 72
        for line in text.splitlines():
            if y < 72:
                pdf.showPage()
                y = height - 72
            pdf.drawString(72, y, line[:120])
            y -= 14
        pdf.save()
        buffer.seek(0)
        return buffer.read()
    except ImportError:
        raise RuntimeError("PDF export requires weasyprint or reportlab to be installed.")
    except Exception as exc:
        raise RuntimeError(f"PDF export failed: {exc}") from exc


# =============================================================================
# EXPORT ORCHESTRATOR
# =============================================================================

class DataExportService:
    """Orchestrates data transformation and export operations."""

    def assess_quality(self, tables: List[EnhancedExtractedTable]) -> DataQualityReport:
        """Assess data quality."""
        return assess_data_quality(tables)

    def clean_data(
        self,
        tables: List[EnhancedExtractedTable],
        operations: List[str] = None,
    ) -> List[EnhancedExtractedTable]:
        """Clean all tables."""
        return [clean_table(t, operations) for t in tables]

    def apply_transformations(
        self,
        table: EnhancedExtractedTable,
        transformations: List[DataTransformation],
    ) -> EnhancedExtractedTable:
        """Apply multiple transformations to a table."""
        result = table
        for transform in transformations:
            result = apply_transformation(result, transform)
        return result

    async def export(
        self,
        result: EnhancedAnalysisResult,
        config: ExportConfiguration,
    ) -> Tuple[bytes, str]:
        """Export analysis result in specified format."""
        filename = config.filename or f"analysis_{result.analysis_id}"

        if config.format == ExportFormat.CSV:
            content = export_to_csv(result.tables)
            return content.encode('utf-8'), f"{filename}.csv"

        elif config.format == ExportFormat.JSON:
            content = export_to_json(result, config.include_raw_data)
            return content.encode('utf-8'), f"{filename}.json"

        elif config.format == ExportFormat.MARKDOWN:
            content = export_to_markdown(result)
            return content.encode('utf-8'), f"{filename}.md"

        elif config.format == ExportFormat.HTML:
            content = export_to_html(result)
            return content.encode('utf-8'), f"{filename}.html"

        elif config.format == ExportFormat.EXCEL:
            content = await export_to_excel(result, config.include_charts)
            return content, f"{filename}.xlsx"

        elif config.format == ExportFormat.PDF:
            content = await export_to_pdf(result)
            return content, f"{filename}.pdf"

        else:
            # Default to JSON
            content = export_to_json(result)
            return content.encode('utf-8'), f"{filename}.json"
