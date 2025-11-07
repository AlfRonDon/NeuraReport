from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Alignment, Border, Font, Side  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore
    Alignment = None  # type: ignore
    Border = None  # type: ignore
    Font = None  # type: ignore
    Side = None  # type: ignore
    get_column_letter = None  # type: ignore
    Table = None  # type: ignore
    TableStyleInfo = None  # type: ignore

from .html_table_parser import extract_first_table

logger = logging.getLogger("neura.reports.xlsx")


def _auto_column_widths(worksheet) -> None:
    if get_column_letter is None:  # pragma: no cover
        return
    for col_idx in range(1, worksheet.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in worksheet[letter]:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            max_len = max(max_len, len(text))
        width = min(120, max(12, max_len + 2))
        worksheet.column_dimensions[letter].width = width


def html_file_to_xlsx(html_path: Path, output_path: Path) -> Optional[Path]:
    if openpyxl is None:  # pragma: no cover
        logger.warning(
            "xlsx_export_unavailable",
            extra={
                "event": "xlsx_export_unavailable",
                "reason": "openpyxl not installed",
                "html_path": str(html_path),
            },
        )
        return None

    html_text = html_path.read_text(encoding="utf-8", errors="ignore")
    rows = extract_first_table(html_text)
    if not rows:
        rows = [[line.strip()] for line in html_text.splitlines() if line.strip()]
        if not rows:
            rows = [["Report output unavailable"]]

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if Alignment is not None:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True,
                )
            if Font is not None and r_idx == 1:
                cell.font = Font(bold=True)

    if Border is not None and Side is not None:
        thin = Side(style="thin", color="FFC0C0C0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        ):
            for cell in row:
                cell.border = border

    if ws.max_row > 1:
        ws.freeze_panes = "A2"
    if Table is not None and TableStyleInfo is not None and ws.max_row > 1 and ws.max_column > 0:
        ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
        table = Table(displayName="ReportTable", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    else:
        ws.auto_filter.ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"

    _auto_column_widths(ws)

    try:
        wb.save(output_path)
    except Exception as exc:  # pragma: no cover
        logger.warning(
            "xlsx_export_save_failed",
            extra={
                "event": "xlsx_export_save_failed",
                "html_path": str(html_path),
                "xlsx_path": str(output_path),
                "error": str(exc),
            },
        )
        return None

    logger.info(
        "xlsx_export_success",
        extra={
            "event": "xlsx_export_success",
            "html_path": str(html_path),
            "xlsx_path": str(output_path),
        },
    )
    return output_path
