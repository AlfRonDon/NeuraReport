from __future__ import annotations

import html
import json
import logging
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

try:
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore

from ..prompts.llm_prompts_excel import build_excel_llm_call_1_prompt
from ..templates.TemplateVerify import MODEL, get_openai_client
from ..utils import call_chat_completion, extract_tokens, normalize_token_braces, strip_code_fences
from ..utils.render import render_html_to_png

logger = logging.getLogger("neura.excel.verify")


@dataclass
class ExcelInitialResult:
    html_path: Path
    png_path: Optional[Path]


def _extract_marked_section(text: str, begin: str, end: str) -> Optional[str]:
    pattern = re.compile(re.escape(begin) + r"([\s\S]*?)" + re.escape(end))
    match = pattern.search(text)
    if not match:
        return None
    return match.group(1).strip()


def _normalize_token(name: str) -> str:
    if name is None:
        return ""
    text = str(name).strip().lower()
    if not text:
        return ""
    text = re.sub(r"[^0-9a-z]+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def _row_has_values(values) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            if value.strip():
                return True
            continue
        return True
    return False


def _ensure_label(value: object, idx: int) -> str:
    if value not in (None, ""):
        text = str(value).strip()
        if text:
            return text
    return f"Column {idx + 1}"


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    try:
        text = str(value)
    except Exception:
        return ""
    return text.strip()


def _build_placeholder_samples(tokens: list[str], data_row: list[str]) -> dict[str, str]:
    samples: dict[str, str] = {}
    if not tokens:
        return samples
    for idx, token in enumerate(tokens):
        placeholder = token
        value = ""
        if data_row is not None and idx < len(data_row):
            cell = data_row[idx]
            if cell is not None:
                value = str(cell).strip()
        samples[placeholder] = value or "NOT_VISIBLE"
    return samples


def _sheet_snapshot_for_llm(sheet, *, max_rows: int = 20, max_preface_rows: int = 6) -> tuple[dict[str, Any], list[dict[str, Any]], list[str], int]:
    rows = list(sheet.iter_rows(values_only=True))
    header_row = None
    header_index = -1
    for idx, row in enumerate(rows):
        if _row_has_values(row):
            header_row = row
            header_index = idx
            break
    if header_row is None:
        header_row = []
        header_index = -1

    header_labels = [_ensure_label(value, idx) for idx, value in enumerate(header_row)]
    preface_rows = rows[: max(header_index, 0)]

    data_rows = []
    if header_index >= 0:
        for row in rows[header_index + 1 :]:
            if _row_has_values(row):
                data_rows.append(row)
    else:
        data_rows = [row for row in rows if _row_has_values(row)]

    data_row_count = len(data_rows)
    sample_rows: list[dict[str, Any]] = []
    for offset, row in enumerate(data_rows[: max_rows]):
        cells = [_stringify_cell(row[idx] if idx < len(row) else "") for idx in range(len(header_labels))]
        sample_rows.append(
            {
                "row_index": header_index + offset + 2 if header_index >= 0 else offset + 1,
                "cells": cells,
            }
        )

    token_plan: list[dict[str, Any]] = []
    for idx, label in enumerate(header_labels):
        norm = _normalize_token(label) or f"col_{idx + 1}"
        token_name = f"row_{norm}"
        sample_value = ""
        if sample_rows and idx < len(sample_rows[0]["cells"]):
            sample_value = sample_rows[0]["cells"][idx]
        token_plan.append(
            {
                "token": token_name,
                "header": label,
                "column_index": idx,
                "sample": sample_value,
            }
        )

    first_data_row = sample_rows[0]["cells"] if sample_rows else []
    grid_preview: list[list[str]] = []
    preview_limit = header_index + 1 + max_rows if header_index >= 0 else max_rows
    for row in rows[:preview_limit]:
        cols = max(len(header_labels), 1)
        grid_preview.append([_stringify_cell(row[idx] if idx < len(row) else "") for idx in range(cols)])

    snapshot = {
        "sheet_title": str(sheet.title or "Sheet1"),
        "preface_rows": [[_stringify_cell(cell) for cell in row] for row in preface_rows[-max_preface_rows:]],
        "headers": header_labels,
        "token_plan": token_plan,
        "sample_rows": sample_rows,
        "grid_preview": grid_preview,
        "sheet_notes": {
            "header_row_index": header_index,
            "data_row_count": data_row_count,
            "column_count": len(header_labels),
            "non_empty_rows": sum(1 for row in rows if _row_has_values(row)),
        },
    }

    return snapshot, token_plan, first_data_row, data_row_count


def _sheet_to_placeholder_html(sheet) -> tuple[str, list[str], list[str]]:
    rows = list(sheet.iter_rows(values_only=True))
    header_row = None
    header_index = -1
    for idx, row in enumerate(rows):
        if _row_has_values(row):
            header_row = row
            header_index = idx
            break

    placeholder_tokens: list[str] = []
    if header_row:
        header_labels = [_ensure_label(value, idx) for idx, value in enumerate(header_row)]
    else:
        header_labels = []

    placeholder_tokens = []
    placeholder_cells: list[str] = []
    data_labels: list[str] = []
    for idx, label in enumerate(header_labels):
        norm = _normalize_token(label) or f"col_{idx + 1}"
        token = f"row_{norm}"
        placeholder_tokens.append(token)
        placeholder_cells.append("<td>{" + token + "}</td>")
        data_labels.append(norm)

    if placeholder_tokens:
        th_cells = [
            f'<th data-label="{html.escape(data_label)}">{html.escape(label)}</th>'
            for data_label, label in zip(data_labels, header_labels)
        ]
        thead_html = f"<thead><tr>{''.join(th_cells)}</tr></thead>"
        tbody_html = f"<tbody><tr>{''.join(placeholder_cells)}</tr></tbody>"
    else:
        thead_html = "<thead><tr></tr></thead>"
        tbody_html = "<tbody><tr></tr></tbody>"

    styles = """
    <style>
      @page { size: A4; margin: 24mm; }
      body { font-family: Arial, sans-serif; }
      table { border-collapse: collapse; width: 100%; }
      td, th { border: 1px solid #999; padding: 6px 8px; vertical-align: top; }
    </style>
    """
    head = f"<head><meta charset='utf-8'>{styles}</head>"
    title = html.escape(str(sheet.title or "Sheet1"))
    caption = f"<caption style='caption-side:top;font-weight:700;text-align:left;margin:6px 0'>{title}</caption>"
    table = f'<table id="data-table">{caption}{thead_html}{tbody_html}</table>'
    html_text = f"<html>{head}<body>{table}</body></html>"

    first_data_row: list[str] = []
    if header_index >= 0:
        for row in rows[header_index + 1 :]:
            if _row_has_values(row):
                first_data_row = [_stringify_cell(row[idx] if idx < len(row) else "") for idx in range(len(header_labels))]
                break

    return html_text, placeholder_tokens, first_data_row


def _request_excel_llm_template(snapshot: dict[str, Any], sheet_html: str, schema_payload: Optional[dict[str, Any]] = None) -> tuple[str, Optional[dict[str, Any]]]:
    snapshot_json = json.dumps(snapshot, ensure_ascii=False, indent=2)
    sheet_html_payload = sheet_html or "<html></html>"
    schema_json = json.dumps(schema_payload or {}, ensure_ascii=False, separators=(",", ":"))
    prompt = build_excel_llm_call_1_prompt(snapshot_json, sheet_html_payload, schema_json)
    content = [{"type": "text", "text": prompt}]
    if not os.getenv("OPENAI_API_KEY"):
        logger.warning(
            "excel_llm_skipped_no_api_key",
            extra={"event": "excel_llm_skipped_no_api_key"},
        )
        return sheet_html, schema_payload
    client = get_openai_client()
    response = call_chat_completion(
        client,
        model=MODEL,
        messages=[{"role": "user", "content": content}],
        description="excel_template_initial_html",
    )
    raw_content = strip_code_fences(response.choices[0].message.content or "")

    html_section = _extract_marked_section(raw_content, "<!--BEGIN_HTML-->", "<!--END_HTML-->")
    if html_section is None:
        raise RuntimeError("Excel LLM response missing HTML markers")
    html_clean = normalize_token_braces(html_section.strip())

    schema_section = _extract_marked_section(raw_content, "<!--BEGIN_SCHEMA_JSON-->", "<!--END_SCHEMA_JSON-->")
    schema_doc = None
    if schema_section:
        try:
            schema_doc = json.loads(schema_section)
        except json.JSONDecodeError:
            logger.warning(
                "excel_llm_schema_parse_failed",
                extra={"event": "excel_llm_schema_parse_failed", "snippet": schema_section[:200]},
            )
    return html_clean, schema_doc


def _sheet_to_reference_html(sheet, *, max_rows: int = 5) -> str:
    """
    Build a data-only HTML snapshot of the original Excel sheet (no placeholders),
    using the first non-empty row as header and up to `max_rows` subsequent data rows.
    This serves as the reference image for fidelity preview and LLM context.
    """
    rows = list(sheet.iter_rows(values_only=True))

    header_row = None
    header_index = -1
    for i, row in enumerate(rows):
        if _row_has_values(row):
            header_row = row
            header_index = i
            break
    if header_row is None:
        header_row = []
        header_index = -1

    header_labels = [_ensure_label(value, idx) for idx, value in enumerate(header_row)]

    styles = """
    <style>
      @page { size: A4; margin: 24mm; }
      body { font-family: Arial, sans-serif; }
      table { border-collapse: collapse; width: 100%; }
      td, th { border: 1px solid #999; padding: 6px 8px; vertical-align: top; }
    </style>
    """
    head = f"<head><meta charset='utf-8'>{styles}</head>"
    title = html.escape(str(sheet.title or "Sheet1"))
    caption = f"<caption style='caption-side:top;font-weight:700;text-align:left;margin:6px 0'>{title}</caption>"

    th_cells = "".join(f"<th>{html.escape(label)}</th>" for label in header_labels)
    thead_html = f"<thead><tr>{th_cells}</tr></thead>"

    # Collect up to max_rows data rows after header
    body_rows: list[str] = []
    if header_index >= 0:
        for row in rows[header_index + 1 : header_index + 1 + max_rows]:
            if not _row_has_values(row):
                continue
            tds = [html.escape("" if v is None else str(v)) for v in row]
            body_rows.append("<tr>" + "".join(f"<td>{v}</td>" for v in tds) + "</tr>")
    if not body_rows:
        body_rows.append("<tr></tr>")
    tbody_html = "<tbody>" + "".join(body_rows) + "</tbody>"

    table = f'<table id="data-table">{caption}{thead_html}{tbody_html}</table>'
    return f"<html>{head}<body>{table}</body></html>"


def xlsx_to_html_preview(
    excel_path: Path,
    out_dir: Path,
    *,
    page_size: str = "A4",
    dpi: int = 144,
    db_path: Path | None = None,
) -> ExcelInitialResult:
    """
    Load the first worksheet of an Excel file and produce:
      - template_p1.html: LLM-generated HTML template (tokens preserved)
      - schema_ext.json: optional schema emitted by the LLM
      - sample_rows.json: literal samples for row_* tokens based on the first data row
      - report_final.png: screenshot of the generated template
      - reference_p1.html/png: literal snapshot of the worksheet data for fidelity preview
    """
    if openpyxl is None:  # pragma: no cover
        raise RuntimeError("openpyxl is required. Install with `pip install openpyxl`.")

    out_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(filename=str(excel_path), data_only=True)
    sheet = wb.active
    # Enforce a simple safety/UX constraint for initial Excel uploads:
    # Limit the number of non-empty data rows to a maximum (default 30).
    # If exceeded, ask the user to delete extra rows and re-upload.
    try:
        max_rows_env = os.getenv("EXCEL_MAX_DATA_ROWS", "30").strip()
        MAX_DATA_ROWS = int(max_rows_env) if max_rows_env else 30
    except Exception:
        MAX_DATA_ROWS = 30

    snapshot, token_plan, first_data_row, data_row_count = _sheet_snapshot_for_llm(sheet, max_rows=MAX_DATA_ROWS)
    sheet_prototype_html, placeholder_tokens, placeholder_first_row = _sheet_to_placeholder_html(sheet)
    if data_row_count > MAX_DATA_ROWS:
        raise RuntimeError(
            f"Excel verification failed: found {data_row_count} data rows; maximum allowed is {MAX_DATA_ROWS}. "
            "Please delete extra rows and upload the file again."
        )

    placeholder_sample_map = _build_placeholder_samples(
        [token for token in placeholder_tokens],
        placeholder_first_row,
    )

    if not os.getenv("OPENAI_API_KEY"):
        logger.warning(
            "excel_llm_skipped_no_api_key",
            extra={"event": "excel_llm_skipped_no_api_key"},
        )
        html_text, schema_payload = sheet_prototype_html, None
    else:
        html_text, schema_payload = _request_excel_llm_template(snapshot, sheet_prototype_html)
        tokens_expected = set(placeholder_tokens)
        tokens_present = set(extract_tokens(normalize_token_braces(html_text)))
        missing_tokens = sorted(tokens_expected - tokens_present)
        if missing_tokens:
            logger.warning(
                "excel_llm_missing_tokens",
                extra={"event": "excel_llm_missing_tokens", "missing": missing_tokens},
            )
            html_text, schema_payload = sheet_prototype_html, None
    html_path = out_dir / "template_p1.html"
    html_path.write_text(html_text, encoding="utf-8")

    if schema_payload:
        schema_path = out_dir / "schema_ext.json"
        schema_path.write_text(json.dumps(schema_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    tokens_in_template = set(extract_tokens(normalize_token_braces(html_text)))
    sample_row_map = {
        token: placeholder_sample_map.get(token, "NOT_VISIBLE") for token in tokens_in_template if token in placeholder_sample_map
    }
    sample_payload = {"sample_row": sample_row_map}
    sample_rows_path = out_dir / "sample_rows.json"
    sample_rows_path.write_text(json.dumps(sample_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    # Build a data-only reference HTML to snapshot the original Excel content
    reference_html = _sheet_to_reference_html(sheet, max_rows=MAX_DATA_ROWS)
    reference_html_path = out_dir / "reference_p1.html"
    reference_html_path.write_text(reference_html, encoding="utf-8")

    reference_png_path: Optional[Path] = None
    try:
        reference_png_path = out_dir / "reference_p1.png"
        render_html_to_png(reference_html_path, reference_png_path, page_size=page_size, dpi=dpi)
    except Exception:
        logger.warning(
            "excel_reference_png_render_failed",
            extra={"event": "excel_reference_png_render_failed", "html": str(reference_html_path)},
            exc_info=True,
        )
        reference_png_path = None

    template_png_path: Optional[Path] = None
    try:
        template_png_path = out_dir / "report_final.png"
        render_html_to_png(html_path, template_png_path, page_size=page_size, dpi=dpi)
    except Exception:
        logger.warning(
            "excel_template_png_render_failed",
            extra={"event": "excel_template_png_render_failed", "html": str(html_path)},
            exc_info=True,
        )
        template_png_path = None

    thumbnail_path = template_png_path or reference_png_path
    return ExcelInitialResult(html_path=html_path, png_path=thumbnail_path)
