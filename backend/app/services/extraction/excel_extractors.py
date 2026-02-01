# mypy: ignore-errors
"""
Excel/Spreadsheet Extraction Module.

Supports:
- openpyxl (.xlsx, .xlsm)
- xlrd (.xls)
- pandas (general purpose)
- CSV/TSV files
- ODS (LibreOffice/OpenOffice)

Features:
- Intelligent format detection
- Streaming extraction for large files
- Data type inference and preservation
- Smart header detection with confidence scoring
- Column statistics and profiling
- Memory-efficient chunk processing
- Multiple encoding support for CSV
- Formula value extraction
"""
from __future__ import annotations

import csv
import io
import logging
import re
import time
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Callable, Dict, Generator, Iterator, List, Optional, Tuple, Union

logger = logging.getLogger("neura.extraction.excel")


# =============================================================================
# Configuration and Types
# =============================================================================

class DataType(Enum):
    """Detected data types for columns."""
    STRING = "string"
    INTEGER = "integer"
    FLOAT = "float"
    DATE = "date"
    DATETIME = "datetime"
    BOOLEAN = "boolean"
    CURRENCY = "currency"
    PERCENTAGE = "percentage"
    EMPTY = "empty"
    MIXED = "mixed"


@dataclass
class ColumnStats:
    """Statistics for a column."""
    name: str
    data_type: DataType
    non_empty_count: int
    empty_count: int
    unique_count: int
    min_value: Optional[Any] = None
    max_value: Optional[Any] = None
    sample_values: List[Any] = field(default_factory=list)

    @property
    def fill_rate(self) -> float:
        """Percentage of non-empty values."""
        total = self.non_empty_count + self.empty_count
        return self.non_empty_count / total if total > 0 else 0.0


@dataclass
class ExtractionConfig:
    """Configuration for Excel extraction."""
    max_rows: int = 50000
    max_sheets: int = 50
    max_columns: int = 500
    detect_headers: bool = True
    infer_types: bool = True
    compute_stats: bool = True
    chunk_size: int = 1000
    encodings_to_try: Tuple[str, ...] = ("utf-8", "latin-1", "cp1252", "iso-8859-1")
    date_formats: Tuple[str, ...] = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
    )


DEFAULT_CONFIG = ExtractionConfig()


@dataclass
class ExcelSheet:
    """Data from a single Excel sheet."""
    name: str
    headers: List[str]
    rows: List[List[Any]]
    row_count: int
    column_count: int
    metadata: Dict[str, Any] = field(default_factory=dict)
    column_stats: List[ColumnStats] = field(default_factory=list)
    header_confidence: float = 1.0

    def __post_init__(self):
        """Validate sheet data."""
        # Ensure headers are strings
        self.headers = [str(h) if h is not None else "" for h in self.headers]

    def get_column(self, col_name: str) -> List[Any]:
        """Get all values from a column by header name."""
        try:
            idx = self.headers.index(col_name)
            return [row[idx] if idx < len(row) else None for row in self.rows]
        except ValueError:
            return []

    def get_column_by_index(self, idx: int) -> List[Any]:
        """Get all values from a column by index."""
        if idx < 0 or idx >= self.column_count:
            return []
        return [row[idx] if idx < len(row) else None for row in self.rows]

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "name": self.name,
            "headers": self.headers,
            "rows": self.rows,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "metadata": self.metadata,
            "header_confidence": self.header_confidence,
            "column_stats": [
                {
                    "name": s.name,
                    "data_type": s.data_type.value,
                    "non_empty_count": s.non_empty_count,
                    "fill_rate": s.fill_rate,
                }
                for s in self.column_stats
            ] if self.column_stats else [],
        }

    def iter_rows(self) -> Iterator[Dict[str, Any]]:
        """Iterate over rows as dictionaries."""
        for row in self.rows:
            yield {
                self.headers[i]: row[i] if i < len(row) else None
                for i in range(len(self.headers))
            }


@dataclass
class ExcelExtractionResult:
    """Result of Excel extraction."""
    sheets: List[ExcelSheet]
    filename: str
    format: str
    errors: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)
    extraction_time_ms: float = 0.0

    @property
    def total_rows(self) -> int:
        """Total rows across all sheets."""
        return sum(s.row_count for s in self.sheets)

    @property
    def has_data(self) -> bool:
        """Check if any data was extracted."""
        return any(s.row_count > 0 for s in self.sheets)

    def get_sheet(self, name: str) -> Optional[ExcelSheet]:
        """Get sheet by name."""
        for sheet in self.sheets:
            if sheet.name == name:
                return sheet
        return None

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "sheets": [s.to_dict() for s in self.sheets],
            "filename": self.filename,
            "format": self.format,
            "errors": self.errors,
            "metadata": self.metadata,
            "extraction_time_ms": self.extraction_time_ms,
            "total_rows": self.total_rows,
            "has_data": self.has_data,
        }


# =============================================================================
# Data Type Detection
# =============================================================================

def _detect_data_type(value: Any) -> DataType:
    """Detect the data type of a value."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return DataType.EMPTY

    if isinstance(value, bool):
        return DataType.BOOLEAN

    if isinstance(value, int):
        return DataType.INTEGER

    if isinstance(value, float):
        return DataType.FLOAT

    if isinstance(value, datetime):
        return DataType.DATETIME

    if not isinstance(value, str):
        return DataType.STRING

    text = value.strip()

    # Check for boolean
    if text.lower() in ("true", "false", "yes", "no", "1", "0"):
        return DataType.BOOLEAN

    # Check for currency
    if re.match(r'^[$€£¥₹]?\s*-?\d{1,3}(,\d{3})*(\.\d{2})?$', text):
        return DataType.CURRENCY

    # Check for percentage
    if re.match(r'^-?\d+(\.\d+)?%$', text):
        return DataType.PERCENTAGE

    # Check for integer
    if re.match(r'^-?\d+$', text):
        return DataType.INTEGER

    # Check for float
    if re.match(r'^-?\d+\.\d+$', text):
        return DataType.FLOAT

    # Check for date patterns
    date_patterns = [
        r'^\d{4}-\d{2}-\d{2}$',
        r'^\d{2}/\d{2}/\d{4}$',
        r'^\d{2}-\d{2}-\d{4}$',
    ]
    for pattern in date_patterns:
        if re.match(pattern, text):
            return DataType.DATE

    # Check for datetime patterns
    if re.match(r'^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}', text):
        return DataType.DATETIME

    return DataType.STRING


def _infer_column_type(values: List[Any]) -> DataType:
    """Infer the predominant data type for a column."""
    type_counts: Dict[DataType, int] = {}

    for value in values:
        dtype = _detect_data_type(value)
        if dtype != DataType.EMPTY:
            type_counts[dtype] = type_counts.get(dtype, 0) + 1

    if not type_counts:
        return DataType.EMPTY

    # Get most common type
    most_common = max(type_counts.items(), key=lambda x: x[1])
    total_non_empty = sum(type_counts.values())

    # If most common type covers >80% of values, use it
    if most_common[1] / total_non_empty >= 0.8:
        return most_common[0]

    return DataType.MIXED


def _compute_column_stats(
    header: str,
    values: List[Any],
    config: ExtractionConfig,
) -> ColumnStats:
    """Compute statistics for a column."""
    non_empty = [v for v in values if v is not None and (not isinstance(v, str) or v.strip())]
    empty_count = len(values) - len(non_empty)

    # Infer type
    data_type = _infer_column_type(values) if config.infer_types else DataType.STRING

    # Calculate unique count (sample if too large)
    sample_for_unique = non_empty[:1000] if len(non_empty) > 1000 else non_empty
    try:
        unique_count = len(set(str(v) for v in sample_for_unique))
    except Exception as e:
        logger.debug(f"Could not compute unique count for column '{header}': {e}")
        unique_count = 0

    # Get min/max for numeric types
    min_val = None
    max_val = None
    if data_type in (DataType.INTEGER, DataType.FLOAT, DataType.CURRENCY):
        numeric_vals = []
        for v in non_empty:
            try:
                if isinstance(v, (int, float)):
                    numeric_vals.append(v)
                elif isinstance(v, str):
                    cleaned = v.replace("$", "").replace(",", "").replace("%", "").strip()
                    numeric_vals.append(float(cleaned))
            except (ValueError, TypeError):
                # Expected for non-numeric values, skip silently
                pass
        if numeric_vals:
            min_val = min(numeric_vals)
            max_val = max(numeric_vals)

    # Sample values
    sample_values = non_empty[:5] if non_empty else []

    return ColumnStats(
        name=header,
        data_type=data_type,
        non_empty_count=len(non_empty),
        empty_count=empty_count,
        unique_count=unique_count,
        min_value=min_val,
        max_value=max_val,
        sample_values=sample_values,
    )


# =============================================================================
# Header Detection
# =============================================================================

def _calculate_header_confidence(row: List[Any], data_rows: List[List[Any]]) -> float:
    """
    Calculate confidence that a row is a header row.

    Uses multiple heuristics:
    - Headers are usually text, not numbers
    - Headers are often shorter than data
    - Headers have distinct patterns
    """
    if not row:
        return 0.0

    confidence = 0.5  # Start neutral

    # Check if row contains mostly text
    text_count = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
    num_count = sum(1 for cell in row if isinstance(cell, (int, float)))
    non_empty = text_count + num_count

    if non_empty == 0:
        return 0.0

    text_ratio = text_count / non_empty
    if text_ratio > 0.8:
        confidence += 0.2
    elif text_ratio < 0.3:
        confidence -= 0.2

    # Check if data rows have different patterns
    if data_rows:
        data_text_ratios = []
        for data_row in data_rows[:10]:
            data_text = sum(1 for cell in data_row if isinstance(cell, str) and cell.strip())
            data_num = sum(1 for cell in data_row if isinstance(cell, (int, float)))
            data_total = data_text + data_num
            if data_total > 0:
                data_text_ratios.append(data_text / data_total)

        if data_text_ratios:
            avg_data_text_ratio = sum(data_text_ratios) / len(data_text_ratios)
            if text_ratio > avg_data_text_ratio + 0.3:
                confidence += 0.2

    # Check for common header patterns
    header_keywords = {
        "id", "name", "date", "time", "amount", "total", "qty", "quantity",
        "price", "description", "status", "type", "category", "email", "phone",
        "address", "city", "country", "code", "number", "#", "no", "count",
    }
    keyword_matches = sum(
        1 for cell in row
        if isinstance(cell, str) and cell.strip().lower() in header_keywords
    )
    if keyword_matches > 0:
        confidence += min(0.2, keyword_matches * 0.05)

    # Check for title case or all caps
    title_or_caps = sum(
        1 for cell in row
        if isinstance(cell, str) and (cell.istitle() or cell.isupper())
    )
    if title_or_caps > len(row) / 2:
        confidence += 0.1

    return min(1.0, max(0.0, confidence))


class ExcelExtractor:
    """
    Excel/spreadsheet data extractor.

    Supports multiple formats and handles:
    - Multiple sheets
    - Intelligent header detection
    - Data type inference and preservation
    - Large file handling with streaming
    - Column statistics
    - Multiple encoding support
    """

    def __init__(
        self,
        config: Optional[ExtractionConfig] = None,
        max_rows: int = 10000,
        max_sheets: int = 20,
        detect_headers: bool = True,
    ):
        self.config = config or ExtractionConfig(
            max_rows=max_rows,
            max_sheets=max_sheets,
            detect_headers=detect_headers,
        )

    def extract(
        self,
        file_path: Union[str, Path],
        sheet_names: Optional[List[str]] = None,
    ) -> ExcelExtractionResult:
        """
        Extract data from an Excel file.

        Args:
            file_path: Path to Excel file
            sheet_names: Optional list of sheet names to extract

        Returns:
            ExcelExtractionResult with all sheet data
        """
        start_time = time.time()
        file_path = Path(file_path)
        suffix = file_path.suffix.lower()

        result: ExcelExtractionResult

        if suffix == ".csv":
            result = self._extract_csv(file_path)
        elif suffix == ".tsv":
            result = self._extract_csv(file_path, delimiter="\t")
        elif suffix in (".xlsx", ".xlsm"):
            result = self._extract_xlsx(file_path, sheet_names)
        elif suffix == ".xls":
            result = self._extract_xls(file_path, sheet_names)
        elif suffix == ".ods":
            result = self._extract_ods(file_path, sheet_names)
        else:
            result = ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="unknown",
                errors=[f"Unsupported format: {suffix}"],
            )

        result.extraction_time_ms = (time.time() - start_time) * 1000
        return result

    def _extract_ods(
        self,
        file_path: Path,
        sheet_names: Optional[List[str]] = None,
    ) -> ExcelExtractionResult:
        """Extract from ODS (LibreOffice/OpenOffice) files."""
        try:
            import pandas as pd
        except ImportError:
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="ods",
                errors=["pandas and odfpy required. Run: pip install pandas odfpy"],
            )

        sheets: List[ExcelSheet] = []
        errors: List[str] = []

        try:
            # Read all sheets
            ods_data = pd.read_excel(file_path, engine="odf", sheet_name=None)

            sheets_to_process = sheet_names or list(ods_data.keys())[:self.config.max_sheets]

            for sheet_name in sheets_to_process:
                if sheet_name not in ods_data:
                    errors.append(f"Sheet '{sheet_name}' not found")
                    continue

                df = ods_data[sheet_name]
                if df.empty:
                    continue

                # Truncate if needed
                if len(df) > self.config.max_rows:
                    errors.append(f"Sheet '{sheet_name}' truncated at {self.config.max_rows} rows")
                    df = df.head(self.config.max_rows)

                # Convert to lists
                headers = [str(col) for col in df.columns.tolist()]
                rows = df.fillna("").values.tolist()

                # Compute stats if enabled
                column_stats = []
                if self.config.compute_stats:
                    for i, header in enumerate(headers):
                        col_values = [row[i] if i < len(row) else None for row in rows]
                        column_stats.append(_compute_column_stats(header, col_values, self.config))

                sheets.append(ExcelSheet(
                    name=sheet_name,
                    headers=headers,
                    rows=rows,
                    row_count=len(rows),
                    column_count=len(headers),
                    column_stats=column_stats,
                ))

            return ExcelExtractionResult(
                sheets=sheets,
                filename=file_path.name,
                format="ods",
                errors=errors,
            )

        except Exception as e:
            logger.error(f"ODS extraction failed: {e}")
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="ods",
                errors=["Extraction failed"],
            )

    def _extract_xlsx(
        self,
        file_path: Path,
        sheet_names: Optional[List[str]] = None,
    ) -> ExcelExtractionResult:
        """Extract from .xlsx/.xlsm files using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="xlsx",
                errors=["openpyxl not installed. Run: pip install openpyxl"],
            )

        sheets: List[ExcelSheet] = []
        errors: List[str] = []

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

            sheets_to_process = sheet_names or workbook.sheetnames[:self.config.max_sheets]

            if len(workbook.sheetnames) > self.config.max_sheets and not sheet_names:
                errors.append(f"File has {len(workbook.sheetnames)} sheets, processing first {self.config.max_sheets}")

            for sheet_name in sheets_to_process:
                if sheet_name not in workbook.sheetnames:
                    errors.append(f"Sheet '{sheet_name}' not found")
                    continue

                sheet = workbook[sheet_name]

                # Read data
                rows_data: List[List[Any]] = []
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    if row_idx >= self.config.max_rows:
                        errors.append(f"Sheet '{sheet_name}' truncated at {self.config.max_rows} rows")
                        break

                    # Convert row to list and clean values
                    cleaned_row = []
                    for cell in row:
                        if cell is None:
                            cleaned_row.append("")
                        elif isinstance(cell, datetime):
                            cleaned_row.append(cell)
                        elif isinstance(cell, (int, float)):
                            cleaned_row.append(cell)
                        else:
                            cleaned_row.append(str(cell))
                    rows_data.append(cleaned_row)

                if not rows_data:
                    continue

                # Calculate header confidence and detect headers
                header_confidence = 1.0
                if self.config.detect_headers and rows_data:
                    header_confidence = _calculate_header_confidence(rows_data[0], rows_data[1:])

                    if header_confidence >= 0.5:
                        headers = [str(h) if h else f"Column_{i+1}" for i, h in enumerate(rows_data[0])]
                        data_rows = rows_data[1:]
                    else:
                        headers = [f"Column_{i+1}" for i in range(len(rows_data[0]))]
                        data_rows = rows_data
                else:
                    headers = [f"Column_{i+1}" for i in range(len(rows_data[0]))]
                    data_rows = rows_data

                # Normalize row lengths
                max_cols = min(len(headers), self.config.max_columns)
                headers = headers[:max_cols]
                normalized_rows = []
                for row in data_rows:
                    normalized = list(row)
                    while len(normalized) < max_cols:
                        normalized.append("")
                    normalized_rows.append(normalized[:max_cols])

                # Compute column statistics if enabled
                column_stats = []
                if self.config.compute_stats and normalized_rows:
                    for i, header in enumerate(headers):
                        col_values = [row[i] if i < len(row) else None for row in normalized_rows]
                        column_stats.append(_compute_column_stats(header, col_values, self.config))

                sheets.append(ExcelSheet(
                    name=sheet_name,
                    headers=headers,
                    rows=normalized_rows,
                    row_count=len(normalized_rows),
                    column_count=len(headers),
                    header_confidence=header_confidence,
                    column_stats=column_stats,
                ))

            workbook.close()

            return ExcelExtractionResult(
                sheets=sheets,
                filename=file_path.name,
                format="xlsx",
                errors=errors,
                metadata={"total_sheets": len(workbook.sheetnames)},
            )

        except Exception as e:
            logger.error(f"XLSX extraction failed: {e}")
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="xlsx",
                errors=["Extraction failed"],
            )

    def _extract_xls(
        self,
        file_path: Path,
        sheet_names: Optional[List[str]] = None,
    ) -> ExcelExtractionResult:
        """Extract from .xls files using xlrd."""
        try:
            import xlrd
        except ImportError:
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="xls",
                errors=["xlrd not installed. Run: pip install xlrd"],
            )

        sheets: List[ExcelSheet] = []
        errors: List[str] = []

        try:
            workbook = xlrd.open_workbook(file_path)

            sheet_list = sheet_names or workbook.sheet_names()[:self.max_sheets]

            for sheet_name in sheet_list:
                try:
                    sheet = workbook.sheet_by_name(sheet_name)
                except xlrd.XLRDError:
                    errors.append(f"Sheet '{sheet_name}' not found")
                    continue

                rows_data: List[List[Any]] = []
                for row_idx in range(min(sheet.nrows, self.max_rows)):
                    row = []
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        if cell.ctype == xlrd.XL_CELL_EMPTY:
                            row.append("")
                        elif cell.ctype == xlrd.XL_CELL_NUMBER:
                            row.append(cell.value)
                        else:
                            row.append(str(cell.value))
                    rows_data.append(row)

                if not rows_data:
                    continue

                # Detect headers
                if self.detect_headers and rows_data:
                    headers = [str(h) if h else f"Column_{i+1}" for i, h in enumerate(rows_data[0])]
                    data_rows = rows_data[1:]
                else:
                    headers = [f"Column_{i+1}" for i in range(len(rows_data[0]))]
                    data_rows = rows_data

                sheets.append(ExcelSheet(
                    name=sheet_name,
                    headers=headers,
                    rows=data_rows,
                    row_count=len(data_rows),
                    column_count=len(headers),
                ))

            return ExcelExtractionResult(
                sheets=sheets,
                filename=file_path.name,
                format="xls",
                errors=errors,
            )

        except Exception as e:
            logger.error(f"XLS extraction failed: {e}")
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="xls",
                errors=["Extraction failed"],
            )

    def _detect_encoding(self, file_path: Path) -> str:
        """Detect file encoding by trying multiple encodings."""
        # Try chardet if available
        try:
            import chardet
            with open(file_path, 'rb') as f:
                raw = f.read(10000)
            result = chardet.detect(raw)
            if result['confidence'] > 0.7:
                return result['encoding']
        except ImportError:
            pass

        # Fallback: try each encoding
        for encoding in self.config.encodings_to_try:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    f.read(1000)
                return encoding
            except (UnicodeDecodeError, LookupError):
                continue

        return 'utf-8'

    def _extract_csv(
        self,
        file_path: Path,
        delimiter: str = ",",
    ) -> ExcelExtractionResult:
        """Extract from CSV/TSV files with smart encoding detection."""
        errors: List[str] = []

        # Detect encoding
        encoding = self._detect_encoding(file_path)

        try:
            with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                # Detect delimiter
                sample = f.read(8192)
                f.seek(0)

                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
                    delimiter = dialect.delimiter
                except csv.Error:
                    pass

                reader = csv.reader(f, delimiter=delimiter)
                rows_data = list(reader)

            if not rows_data:
                return ExcelExtractionResult(
                    sheets=[],
                    filename=file_path.name,
                    format="csv",
                    errors=["CSV file is empty"],
                )

            # Truncate if needed
            if len(rows_data) > self.config.max_rows:
                errors.append(f"CSV truncated at {self.config.max_rows} rows")
                rows_data = rows_data[:self.config.max_rows]

            # Calculate header confidence
            header_confidence = 1.0
            if self.config.detect_headers and rows_data:
                header_confidence = _calculate_header_confidence(rows_data[0], rows_data[1:])

                if header_confidence >= 0.5:
                    headers = [str(h) if h else f"Column_{i+1}" for i, h in enumerate(rows_data[0])]
                    data_rows = rows_data[1:]
                else:
                    max_cols = max(len(row) for row in rows_data) if rows_data else 0
                    headers = [f"Column_{i+1}" for i in range(max_cols)]
                    data_rows = rows_data
            else:
                max_cols = max(len(row) for row in rows_data) if rows_data else 0
                headers = [f"Column_{i+1}" for i in range(max_cols)]
                data_rows = rows_data

            # Normalize row lengths
            max_cols = min(len(headers), self.config.max_columns)
            headers = headers[:max_cols]
            normalized_rows = []
            for row in data_rows:
                normalized = list(row)
                while len(normalized) < max_cols:
                    normalized.append("")
                normalized_rows.append(normalized[:max_cols])

            # Compute column statistics if enabled
            column_stats = []
            if self.config.compute_stats and normalized_rows:
                for i, header in enumerate(headers):
                    col_values = [row[i] if i < len(row) else None for row in normalized_rows]
                    column_stats.append(_compute_column_stats(header, col_values, self.config))

            sheets = [ExcelSheet(
                name=file_path.stem,
                headers=headers,
                rows=normalized_rows,
                row_count=len(normalized_rows),
                column_count=len(headers),
                header_confidence=header_confidence,
                column_stats=column_stats,
                metadata={"encoding": encoding, "delimiter": delimiter},
            )]

            return ExcelExtractionResult(
                sheets=sheets,
                filename=file_path.name,
                format="csv",
                errors=errors,
                metadata={"encoding": encoding},
            )

        except Exception as e:
            logger.error(f"CSV extraction failed: {e}")
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="csv",
                errors=["Extraction failed"],
            )


def extract_excel_data(
    file_path: Union[str, Path],
    sheet_names: Optional[List[str]] = None,
    max_rows: int = 10000,
) -> ExcelExtractionResult:
    """
    Quick function to extract data from Excel/CSV files.

    Args:
        file_path: Path to file
        sheet_names: Optional list of sheets to extract
        max_rows: Maximum rows per sheet

    Returns:
        ExcelExtractionResult
    """
    extractor = ExcelExtractor(max_rows=max_rows)
    return extractor.extract(file_path, sheet_names)
