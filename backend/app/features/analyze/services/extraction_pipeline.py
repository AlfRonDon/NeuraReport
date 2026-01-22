# mypy: ignore-errors
"""Extraction pipeline for PDF and Excel documents."""
from __future__ import annotations

import logging
import re
import time
from dataclasses import dataclass, field
import os
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger("neura.analyze.extraction")

try:
    import fitz  # PyMuPDF
except ImportError:  # pragma: no cover
    fitz = None  # type: ignore

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore

_DEFAULT_MAX_PDF_PAGES = int(os.getenv("NEURA_ANALYSIS_PDF_MAX_PAGES", "50"))
_DEFAULT_MAX_PDF_SECONDS = int(os.getenv("NEURA_ANALYSIS_PDF_MAX_SECONDS", "20"))
_DEFAULT_MAX_EXCEL_ROWS = int(os.getenv("NEURA_ANALYSIS_EXCEL_MAX_ROWS", "500"))
_DEFAULT_MAX_EXCEL_SHEETS = int(os.getenv("NEURA_ANALYSIS_EXCEL_MAX_SHEETS", "10"))


@dataclass
class ExtractedContent:
    """Content extracted from a document before LLM processing."""

    document_type: str  # "pdf" | "excel"
    file_name: str
    page_count: int = 1
    text_content: str = ""
    tables_raw: list[dict[str, Any]] = field(default_factory=list)
    images: list[bytes] = field(default_factory=list)
    sheets: list[dict[str, Any]] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def _stringify_cell(value: object) -> str:
    """Convert cell value to string."""
    if value is None:
        return ""
    try:
        text = str(value)
    except Exception as exc:
        logger.debug(f"Failed to stringify cell value: {exc}")
        return ""
    return text.strip()


def _row_has_values(values) -> bool:
    """Check if a row has any non-empty values."""
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            if value.strip():
                return True
            continue
        return True
    return False


def _ensure_label(value: object, idx: int) -> str:
    """Ensure a column has a label."""
    if value not in (None, ""):
        text = str(value).strip()
        if text:
            return text
    return f"Column {idx + 1}"


def _infer_data_type_from_values(values: list[str]) -> str:
    """Infer data type from a list of string values."""
    if not values:
        return "text"

    date_patterns = [
        r"^\d{4}-\d{2}-\d{2}",
        r"^\d{2}/\d{2}/\d{4}",
        r"^\d{2}-\d{2}-\d{4}",
        r"^\d{1,2}/\d{1,2}/\d{2,4}",
    ]

    numeric_count = 0
    date_count = 0
    total_valid = 0

    for val in values[:30]:
        if not val or not val.strip():
            continue
        total_valid += 1
        str_val = val.strip()

        for pattern in date_patterns:
            if re.match(pattern, str_val):
                date_count += 1
                break
        else:
            try:
                cleaned = str_val.replace(",", "").replace("$", "").replace("%", "").replace(" ", "")
                float(cleaned)
                numeric_count += 1
            except (ValueError, TypeError):
                pass

    if total_valid == 0:
        return "text"
    if date_count >= total_valid * 0.7:
        return "datetime"
    if numeric_count >= total_valid * 0.7:
        return "numeric"
    return "text"


def extract_pdf_content(
    file_path: Path | str,
    file_bytes: bytes | None = None,
    max_pages: int | None = None,
    max_seconds: int | None = None,
) -> ExtractedContent:
    """Extract text, tables, and images from a PDF file."""
    if fitz is None:
        return ExtractedContent(
            document_type="pdf",
            file_name=str(file_path),
            errors=["PyMuPDF (fitz) is not installed. Cannot extract PDF content."],
        )

    file_name = Path(file_path).name if file_path else "document.pdf"

    try:
        if file_bytes:
            doc = fitz.open(stream=file_bytes, filetype="pdf")
        else:
            doc = fitz.open(str(file_path))
    except Exception as exc:
        return ExtractedContent(
            document_type="pdf",
            file_name=file_name,
            errors=[f"Failed to open PDF: {exc}"],
        )

    page_count = len(doc)
    max_pages = max_pages if max_pages is not None else _DEFAULT_MAX_PDF_PAGES
    max_seconds = max_seconds if max_seconds is not None else _DEFAULT_MAX_PDF_SECONDS
    text_parts: list[str] = []
    tables_raw: list[dict[str, Any]] = []
    images: list[bytes] = []
    errors: list[str] = []
    started = time.time()

    if max_pages and page_count > max_pages:
        errors.append(f"PDF has {page_count} pages; processed first {max_pages} pages")

    for page_num, page in enumerate(doc):
        if max_pages and page_num >= max_pages:
            break
        if max_seconds and (time.time() - started) > max_seconds:
            errors.append(f"PDF extraction exceeded time budget of {max_seconds}s; stopping early.")
            break
        try:
            text = page.get_text("text")
            if text.strip():
                text_parts.append(f"--- Page {page_num + 1} ---\n{text}")
        except Exception as exc:
            errors.append(f"Failed to extract text from page {page_num + 1}: {exc}")

        try:
            page_tables = page.find_tables()
            if page_tables and page_tables.tables:
                for table_idx, table in enumerate(page_tables.tables):
                    try:
                        extracted = table.extract()
                        if extracted and len(extracted) > 0:
                            headers = [_stringify_cell(c) or f"Col{i+1}" for i, c in enumerate(extracted[0])]
                            num_cols = len(headers)
                            rows = []
                            for row in extracted[1:]:
                                # Normalize row to match header length
                                normalized_row = [
                                    _stringify_cell(row[i] if i < len(row) else "")
                                    for i in range(num_cols)
                                ]
                                rows.append(normalized_row)

                            col_values: dict[int, list[str]] = {i: [] for i in range(len(headers))}
                            for row in rows[:30]:
                                for i, cell in enumerate(row):
                                    if i < len(headers):
                                        col_values[i].append(cell)

                            data_types = [_infer_data_type_from_values(col_values.get(i, [])) for i in range(len(headers))]

                            tables_raw.append({
                                "id": f"table_p{page_num + 1}_{table_idx + 1}",
                                "headers": headers,
                                "rows": rows,
                                "data_types": data_types,
                                "source_page": page_num + 1,
                            })
                    except Exception as table_exc:
                        errors.append(f"Failed to extract table {table_idx + 1} from page {page_num + 1}: {table_exc}")
        except Exception as exc:
            logger.debug(f"Table extraction not available for page {page_num + 1}: {exc}")

        try:
            image_list = page.get_images(full=True)
            for img_idx, img_info in enumerate(image_list[:3]):
                try:
                    xref = img_info[0]
                    pix = fitz.Pixmap(doc, xref)
                    if pix.n > 4:
                        pix = fitz.Pixmap(fitz.csRGB, pix)
                    img_bytes = pix.tobytes("png")
                    images.append(img_bytes)
                except Exception as img_exc:
                    logger.debug(f"Failed to extract image {img_idx + 1} from page {page_num + 1}: {img_exc}")
        except Exception as img_list_exc:
            logger.debug(f"Failed to get image list from page {page_num + 1}: {img_list_exc}")

    doc.close()

    return ExtractedContent(
        document_type="pdf",
        file_name=file_name,
        page_count=page_count,
        text_content="\n\n".join(text_parts),
        tables_raw=tables_raw,
        images=images[:10],
        errors=errors,
    )


def extract_excel_content(
    file_path: Path | str,
    file_bytes: bytes | None = None,
    max_rows: int | None = None,
    max_sheets: int | None = None,
) -> ExtractedContent:
    """Extract tables and data from an Excel file."""
    if openpyxl is None:
        return ExtractedContent(
            document_type="excel",
            file_name=str(file_path),
            errors=["openpyxl is not installed. Cannot extract Excel content."],
        )

    file_name = Path(file_path).name if file_path else "document.xlsx"

    try:
        if file_bytes:
            wb = openpyxl.load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
        else:
            wb = openpyxl.load_workbook(filename=str(file_path), data_only=True, read_only=True)
    except Exception as exc:
        return ExtractedContent(
            document_type="excel",
            file_name=file_name,
            errors=[f"Failed to open Excel file: {exc}"],
        )

    sheet_count = len(wb.sheetnames)
    tables_raw: list[dict[str, Any]] = []
    sheets_info: list[dict[str, Any]] = []
    text_parts: list[str] = []
    errors: list[str] = []

    max_sheets = max_sheets if max_sheets is not None else _DEFAULT_MAX_EXCEL_SHEETS
    if sheet_count > max_sheets:
        logger.warning(f"Excel file has {sheet_count} sheets, processing only first {max_sheets}")
        errors.append(f"File has {sheet_count} sheets - only the first {max_sheets} were processed")

    for sheet_idx, sheet_name in enumerate(wb.sheetnames[:max_sheets]):
        try:
            sheet = wb[sheet_name]
            max_rows = max_rows if max_rows is not None else _DEFAULT_MAX_EXCEL_ROWS
            header_row = None
            header_index = -1
            headers: list[str] = []
            data_rows: list[list[str]] = []

            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if header_row is None:
                    if _row_has_values(row):
                        header_row = row
                        header_index = idx
                        headers = [_ensure_label(v, i) for i, v in enumerate(header_row)]
                    continue

                if not _row_has_values(row):
                    continue
                if len(data_rows) < max_rows:
                    data_rows.append(
                        [_stringify_cell(row[i] if i < len(row) else "") for i in range(len(headers))]
                    )
                if len(data_rows) >= max_rows:
                    break

            if not data_rows:
                continue

            col_values: dict[int, list[str]] = {i: [] for i in range(len(headers))}
            for row in data_rows[:30]:
                for i, cell in enumerate(row):
                    if i < len(headers):
                        col_values[i].append(cell)

            data_types = [_infer_data_type_from_values(col_values.get(i, [])) for i in range(len(headers))]

            total_rows = max(0, (sheet.max_row or 0) - (header_index + 1))
            truncated = total_rows > max_rows if max_rows else False
            if truncated:
                logger.info(f"Sheet '{sheet_name}' truncated from {total_rows} to {max_rows} rows")

            table_id = f"table_sheet_{sheet_idx + 1}"
            tables_raw.append({
                "id": table_id,
                "title": sheet_name,
                "headers": headers,
                "rows": data_rows[:max_rows],
                "data_types": data_types,
                "source_sheet": sheet_name,
                "truncated": truncated,
                "total_row_count": total_rows,
            })

            sheets_info.append({
                "name": sheet_name,
                "row_count": total_rows,
                "column_count": len(headers),
                "headers": headers,
                "truncated": truncated,
            })

            preview_rows = min(5, len(data_rows))
            text_preview = f"--- Sheet: {sheet_name} ---\n"
            text_preview += f"Headers: {', '.join(headers)}\n"
            text_preview += f"Rows: {total_rows}\n"
            for i in range(preview_rows):
                text_preview += f"Row {i+1}: {', '.join(data_rows[i][:10])}\n"
            text_parts.append(text_preview)

        except Exception as exc:
            errors.append(f"Failed to process sheet '{sheet_name}': {exc}")

    wb.close()

    return ExtractedContent(
        document_type="excel",
        file_name=file_name,
        page_count=sheet_count,
        text_content="\n\n".join(text_parts),
        tables_raw=tables_raw,
        sheets=sheets_info,
        errors=errors,
    )


def extract_document_content(
    file_path: Path | str | None = None,
    file_bytes: bytes | None = None,
    file_name: str | None = None,
    max_pages: int | None = None,
    max_rows: int | None = None,
    max_seconds: int | None = None,
) -> ExtractedContent:
    """Extract content from a document, auto-detecting type from extension."""
    if file_name is None and file_path:
        file_name = Path(file_path).name

    if not file_name:
        file_name = "document"

    ext = Path(file_name).suffix.lower()

    if ext == ".pdf":
        return extract_pdf_content(
            file_path or file_name,
            file_bytes,
            max_pages=max_pages,
            max_seconds=max_seconds,
        )
    elif ext in (".xlsx", ".xls", ".xlsm"):
        return extract_excel_content(file_path or file_name, file_bytes, max_rows=max_rows)
    else:
        return ExtractedContent(
            document_type="unknown",
            file_name=file_name,
            errors=[f"Unsupported file type: {ext}. Only PDF and Excel files are supported."],
        )


def format_content_for_llm(content: ExtractedContent, max_chars: int = 50000) -> str:
    """Format extracted content into a string for LLM processing."""
    parts: list[str] = []

    if content.text_content:
        text_preview = content.text_content[:max_chars // 2]
        parts.append(f"TEXT CONTENT:\n{text_preview}")

    if content.tables_raw:
        parts.append(f"\nEXTRACTED TABLES ({len(content.tables_raw)} found):")
        for table in content.tables_raw[:10]:
            table_str = f"\n[Table: {table.get('id', 'unknown')}]"
            if table.get("title"):
                table_str += f" Title: {table['title']}"
            if table.get("source_page"):
                table_str += f" (Page {table['source_page']})"
            if table.get("source_sheet"):
                table_str += f" (Sheet: {table['source_sheet']})"

            headers = table.get("headers", [])
            table_str += f"\nHeaders: {', '.join(headers)}"

            rows = table.get("rows", [])
            table_str += f"\nRow count: {len(rows)}"

            for i, row in enumerate(rows[:5]):
                row_preview = [str(c)[:50] for c in row[:8]]
                table_str += f"\n  Row {i+1}: {' | '.join(row_preview)}"

            if len(rows) > 5:
                table_str += f"\n  ... and {len(rows) - 5} more rows"

            parts.append(table_str)

    if content.sheets:
        parts.append(f"\nSHEET SUMMARY ({len(content.sheets)} sheets):")
        for sheet in content.sheets:
            parts.append(f"  - {sheet['name']}: {sheet['row_count']} rows, {sheet['column_count']} columns")

    result = "\n".join(parts)
    if len(result) > max_chars:
        result = result[:max_chars] + "\n... (truncated)"

    return result


__all__ = [
    "ExtractedContent",
    "extract_pdf_content",
    "extract_excel_content",
    "extract_document_content",
    "format_content_for_llm",
]
