"""Excel extraction adapter."""

from __future__ import annotations

import logging
import time
from pathlib import Path
from typing import Any, Dict, List

from .base import BaseExtractor, ExtractionResult, ExtractedTable

logger = logging.getLogger("neura.adapters.extraction.excel")


class ExcelExtractor(BaseExtractor):
    """Extract data from Excel documents."""

    SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".xlsb"}

    def supports(self, path: Path) -> bool:
        """Check if this extractor supports the file."""
        return path.suffix.lower() in self.SUPPORTED_EXTENSIONS

    def extract(self, path: Path) -> ExtractionResult:
        """Extract all data from an Excel file."""
        self._validate_path(path)
        start = time.perf_counter()
        errors: List[str] = []
        tables: List[ExtractedTable] = []
        metadata: Dict[str, Any] = {}

        try:
            tables = self.extract_tables(path)
            metadata = self._get_workbook_metadata(path)
        except Exception as e:
            errors.append(f"Excel extraction failed: {e}")
            logger.exception("excel_extraction_failed")

        extraction_time = (time.perf_counter() - start) * 1000

        return ExtractionResult(
            source_path=path,
            page_count=len(tables),  # Sheets as pages
            tables=tables,
            text_blocks=[],
            metadata=metadata,
            errors=errors,
            extraction_time_ms=extraction_time,
        )

    def extract_tables(self, path: Path) -> List[ExtractedTable]:
        """Extract tables from each sheet in Excel file."""
        self._validate_path(path)

        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required. Install with: pip install openpyxl"
            )

        tables: List[ExtractedTable] = []
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]
            table = self._extract_sheet_table(sheet, sheet_idx, sheet_name)
            if table:
                tables.append(table)

        wb.close()
        return tables

    def _extract_sheet_table(
        self,
        sheet,
        sheet_idx: int,
        sheet_name: str,
    ) -> ExtractedTable | None:
        """Extract table data from a single sheet."""
        rows_data: List[List[Any]] = []

        for row in sheet.iter_rows(values_only=True):
            # Convert row to list, handling None values
            row_values = [self._cell_to_string(cell) for cell in row]
            # Skip completely empty rows
            if any(v for v in row_values):
                rows_data.append(row_values)

        if not rows_data:
            return None

        # First non-empty row as headers
        headers = rows_data[0]
        rows = rows_data[1:] if len(rows_data) > 1 else []

        return ExtractedTable(
            page_number=sheet_idx + 1,
            table_index=0,
            headers=headers,
            rows=rows,
            confidence=0.9,
            metadata={"sheet_name": sheet_name},
        )

    def _cell_to_string(self, cell: Any) -> str:
        """Convert cell value to string."""
        if cell is None:
            return ""
        if isinstance(cell, (int, float)):
            # Format numbers nicely
            if isinstance(cell, float) and cell.is_integer():
                return str(int(cell))
            return str(cell)
        return str(cell)

    def _get_workbook_metadata(self, path: Path) -> Dict[str, Any]:
        """Get workbook metadata."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True)
            metadata = {
                "sheet_names": wb.sheetnames,
                "sheet_count": len(wb.sheetnames),
            }
            if wb.properties:
                metadata.update({
                    "title": wb.properties.title,
                    "creator": wb.properties.creator,
                    "created": str(wb.properties.created) if wb.properties.created else None,
                    "modified": str(wb.properties.modified) if wb.properties.modified else None,
                })
            wb.close()
            return metadata
        except Exception:
            return {}
